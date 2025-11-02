#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

xl_path = 'CONTABILIDADE_FINAL_20251029.xlsx'

print("="*80)
print("📊 ANÁLISE DA ABA CAIXA")
print("="*80)

# Primeiro, ler com pandas para ver os valores
print("\n1. VALORES NA ABA CAIXA (Pandas):")
print("-" * 80)

try:
    df_caixa = pd.read_excel(xl_path, sheet_name='CAIXA', header=None, nrows=50)

    # Mostrar primeiras linhas
    for i in range(min(50, len(df_caixa))):
        row = df_caixa.iloc[i]
        # Mostrar apenas colunas não vazias
        valores = []
        for j, val in enumerate(row):
            if pd.notna(val):
                valores.append(f"Col{j}: {val}")
        if valores:
            print(f"Linha {i}: {' | '.join(valores[:6])}")
            if len(valores) > 6:
                print(f"        {' | '.join(valores[6:])}")
except Exception as e:
    print(f"Erro ao ler CAIXA com pandas: {e}")

print("\n" + "="*80)
print("2. FÓRMULAS NA ABA CAIXA (openpyxl):")
print("-" * 80)

try:
    wb = load_workbook(xl_path, data_only=False)

    if 'CAIXA' in wb.sheetnames:
        ws = wb['CAIXA']

        print(f"\nDimensões da aba: {ws.max_row} linhas x {ws.max_column} colunas")

        # Procurar células com fórmulas relacionadas a despesas fixas
        print("\n📝 Células com fórmulas interessantes:")
        print("-" * 80)

        formulas_encontradas = []

        for row in range(1, min(100, ws.max_row + 1)):
            for col in range(1, min(20, ws.max_column + 1)):
                cell = ws.cell(row, col)

                if cell.value and isinstance(cell.value, str):
                    valor_lower = str(cell.value).lower()

                    # Procurar por termos relevantes
                    termos_interesse = ['despesa', 'fixa', 'mensal', 'ordenado', 'alimentação',
                                       'bruno', 'rafael', 'saldo', 'total', 'soma']

                    if any(termo in valor_lower for termo in termos_interesse):
                        coord = f"{get_column_letter(col)}{row}"

                        # Se for fórmula, mostrar a fórmula
                        if str(cell.value).startswith('='):
                            formulas_encontradas.append({
                                'coord': coord,
                                'formula': cell.value,
                                'valor': cell.value
                            })
                        else:
                            # Se não for fórmula mas tem termo interessante, mostrar também
                            if len(str(cell.value)) < 100:
                                print(f"\n{coord}: {cell.value} (texto/valor)")

        # Mostrar fórmulas encontradas
        if formulas_encontradas:
            print("\n🔢 FÓRMULAS ENCONTRADAS:")
            print("-" * 80)
            for item in formulas_encontradas[:20]:  # Limitar a 20 primeiras
                print(f"\n{item['coord']}:")
                print(f"   {item['formula'][:200]}")
                if len(item['formula']) > 200:
                    print(f"   ... (fórmula truncada)")

        # Procurar especificamente por valores relacionados a despesas fixas
        print("\n\n💰 PROCURANDO VALORES ESPECÍFICOS:")
        print("-" * 80)

        # Procurar por valores próximos a €7,939.66 ou €3,969.83
        valores_procurar = [7939.66, 3969.83, 7826.01, 3913.01]

        for row in range(1, min(100, ws.max_row + 1)):
            for col in range(1, min(20, ws.max_column + 1)):
                cell = ws.cell(row, col)

                if isinstance(cell.value, (int, float)):
                    for valor_ref in valores_procurar:
                        if abs(float(cell.value) - valor_ref) < 1:  # Diferença menor que 1€
                            coord = f"{get_column_letter(col)}{row}"

                            # Verificar se célula tem fórmula
                            if hasattr(cell, 'value') and str(ws.cell(row, col).value).startswith('='):
                                print(f"\n{coord}: €{cell.value:.2f}")
                                print(f"   Fórmula: {ws.cell(row, col).value[:200]}")
                            else:
                                print(f"\n{coord}: €{cell.value:.2f} (valor direto)")

                            # Mostrar células ao redor para contexto
                            if col > 1:
                                label_cell = ws.cell(row, col-1)
                                if label_cell.value:
                                    print(f"   Label: {label_cell.value}")
    else:
        print("Aba 'CAIXA' não encontrada!")
        print(f"Abas disponíveis: {wb.sheetnames}")

except Exception as e:
    print(f"Erro ao ler CAIXA com openpyxl: {e}")
    import traceback
    traceback.print_exc()

print("\n" + "="*80)
