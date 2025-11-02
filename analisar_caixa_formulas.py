# -*- coding: utf-8 -*-
"""
Analisar a sheet CAIXA do Excel para entender a lógica de saldos
"""
import pandas as pd
import openpyxl

try:
    print("=" * 100)
    print("ANÁLISE DA SHEET CAIXA - LÓGICA DE SALDOS")
    print("=" * 100)

    # Abrir com openpyxl para ver fórmulas
    wb = openpyxl.load_workbook('CONTABILIDADE_FINAL_20251102.xlsx', data_only=False)
    ws = wb['CAIXA']

    print("\n📊 Primeiras 30 linhas da sheet CAIXA (com fórmulas):")
    print("-" * 100)

    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=30, values_only=False), 1):
        values = []
        for cell in row[:15]:  # Primeiras 15 colunas
            if cell.value is not None:
                # Se for fórmula, mostrar a fórmula
                if isinstance(cell.value, str) and cell.value.startswith('='):
                    values.append(f"[FÓRMULA: {cell.value[:50]}]")
                else:
                    values.append(str(cell.value)[:30])

        if values:
            print(f"Linha {row_idx:2d}: {' | '.join(values)}")

    # Procurar células com "BRUNO" ou "RAFAEL" ou "SALDO"
    print("\n\n" + "=" * 100)
    print("PROCURANDO CÉLULAS COM SALDOS DOS SÓCIOS")
    print("=" * 100)

    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=100, values_only=False), 1):
        for col_idx, cell in enumerate(row[:20], 1):
            if cell.value and isinstance(cell.value, str):
                cell_str = str(cell.value).upper()
                if ('SALDO' in cell_str or 'BRUNO' in cell_str or 'RAFAEL' in cell_str) and \
                   ('BA' in cell_str or 'RR' in cell_str or 'TOTAL' in cell_str):
                    # Mostrar a célula e as próximas 5 colunas
                    print(f"\n📍 Linha {row_idx}, Coluna {col_idx}: {cell.value}")

                    # Mostrar valores/fórmulas das próximas colunas
                    for offset in range(1, 6):
                        next_cell = ws.cell(row=row_idx, column=col_idx + offset)
                        if next_cell.value:
                            if isinstance(next_cell.value, str) and next_cell.value.startswith('='):
                                print(f"   Col {col_idx + offset}: FÓRMULA = {next_cell.value}")
                            else:
                                print(f"   Col {col_idx + offset}: {next_cell.value}")

except Exception as e:
    print(f"❌ Erro: {e}")
    import traceback
    traceback.print_exc()
