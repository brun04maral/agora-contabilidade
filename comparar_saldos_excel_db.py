# -*- coding: utf-8 -*-
"""
Comparar saldos do Excel vs DB
"""
import pandas as pd
import openpyxl
from dotenv import load_dotenv
import os
from sqlalchemy import create_engine, func
from sqlalchemy.orm import sessionmaker
from database.models import Projeto, TipoProjeto, EstadoProjeto, Despesa, TipoDespesa, EstadoDespesa, Boletim, Socio, EstadoBoletim
from decimal import Decimal

load_dotenv()

print("=" * 100)
print("COMPARAÇÃO: EXCEL vs BASE DE DADOS")
print("=" * 100)

# 1. LER VALORES DO EXCEL
print("\n📊 VALORES DO EXCEL (sheet CAIXA):")
print("-" * 100)

wb = openpyxl.load_workbook('CONTABILIDADE_FINAL_20251102.xlsx', data_only=True)
ws = wb['CAIXA']

# Linha 4 = Bruno, Linha 5 = Rafael
bruno_excel = {
    'investimento': ws.cell(row=4, column=3).value or 0,  # C4
    'salarios_atraso': ws.cell(row=4, column=4).value or 0,  # D4
    'projetos_pessoais': ws.cell(row=4, column=5).value or 0,  # E4
    'premios': ws.cell(row=4, column=6).value or 0,  # F4
    'total_ins': ws.cell(row=4, column=7).value or 0,  # G4
    'fixas': ws.cell(row=4, column=8).value or 0,  # H4
    'pessoais_pagas': ws.cell(row=4, column=9).value or 0,  # I4
    'boletins_por_pagar': ws.cell(row=4, column=10).value or 0,  # J4
    'total_outs': ws.cell(row=4, column=11).value or 0,  # K4
    'saldo_com_inv': ws.cell(row=4, column=12).value or 0,  # L4
    'saldo_sem_inv': ws.cell(row=4, column=13).value or 0,  # M4
}

rafael_excel = {
    'investimento': ws.cell(row=5, column=3).value or 0,  # C5
    'salarios_atraso': ws.cell(row=5, column=4).value or 0,  # D5
    'projetos_pessoais': ws.cell(row=5, column=5).value or 0,  # E5
    'premios': ws.cell(row=5, column=6).value or 0,  # F5
    'total_ins': ws.cell(row=5, column=7).value or 0,  # G5
    'fixas': ws.cell(row=5, column=8).value or 0,  # H5
    'pessoais_pagas': ws.cell(row=5, column=9).value or 0,  # I5
    'boletins_por_pagar': ws.cell(row=5, column=10).value or 0,  # J5
    'total_outs': ws.cell(row=5, column=11).value or 0,  # K5
    'saldo_com_inv': ws.cell(row=5, column=12).value or 0,  # L5
    'saldo_sem_inv': ws.cell(row=5, column=13).value or 0,  # M5
}

print("\n👤 BRUNO (Excel):")
for key, val in bruno_excel.items():
    print(f"   {key:20s}: €{float(val):>12,.2f}")

print("\n👤 RAFAEL (Excel):")
for key, val in rafael_excel.items():
    print(f"   {key:20s}: €{float(val):>12,.2f}")

# 2. CALCULAR NA DB
print("\n\n💾 VALORES DA BASE DE DADOS:")
print("-" * 100)

database_url = os.getenv("DATABASE_URL", "sqlite:///./agora_media.db")
engine = create_engine(database_url)
Session = sessionmaker(bind=engine)
db = Session()

try:
    for socio_nome, socio_enum in [("BRUNO", Socio.BRUNO), ("RAFAEL", Socio.RAFAEL)]:
        print(f"\n👤 {socio_nome}:")

        # Projetos pessoais RECEBIDOS
        if socio_nome == "BRUNO":
            projetos_pessoais = db.query(func.sum(Projeto.valor_sem_iva)).filter(
                Projeto.tipo == TipoProjeto.PESSOAL_BRUNO,
                Projeto.estado == EstadoProjeto.RECEBIDO
            ).scalar() or Decimal(0)
        else:
            projetos_pessoais = db.query(func.sum(Projeto.valor_sem_iva)).filter(
                Projeto.tipo == TipoProjeto.PESSOAL_RAFAEL,
                Projeto.estado == EstadoProjeto.RECEBIDO
            ).scalar() or Decimal(0)

        # TODOS os prémios (não filtrar por estado!)
        if socio_nome == "BRUNO":
            premios = db.query(func.sum(Projeto.premio_bruno)).filter(
                Projeto.premio_bruno > 0
            ).scalar() or Decimal(0)
        else:
            premios = db.query(func.sum(Projeto.premio_rafael)).filter(
                Projeto.premio_rafael > 0
            ).scalar() or Decimal(0)

        # Despesas fixas mensais PAGAS
        despesas_fixas = db.query(func.sum(Despesa.valor_sem_iva)).filter(
            Despesa.tipo == TipoDespesa.FIXA_MENSAL,
            Despesa.estado == EstadoDespesa.PAGO
        ).scalar() or Decimal(0)
        despesas_fixas_metade = despesas_fixas / 2

        # Despesas pessoais PAGAS (inclui boletins!)
        if socio_nome == "BRUNO":
            despesas_pessoais_pagas = db.query(func.sum(Despesa.valor_sem_iva)).filter(
                Despesa.tipo == TipoDespesa.PESSOAL_BRUNO,
                Despesa.estado == EstadoDespesa.PAGO
            ).scalar() or Decimal(0)
        else:
            despesas_pessoais_pagas = db.query(func.sum(Despesa.valor_sem_iva)).filter(
                Despesa.tipo == TipoDespesa.PESSOAL_RAFAEL,
                Despesa.estado == EstadoDespesa.PAGO
            ).scalar() or Decimal(0)

        # Boletins PAGOS (para verificação)
        boletins_pagos = db.query(func.sum(Boletim.valor)).filter(
            Boletim.socio == socio_enum,
            Boletim.estado == EstadoBoletim.PAGO
        ).scalar() or Decimal(0)

        total_ins = projetos_pessoais + premios
        total_outs = despesas_fixas_metade + despesas_pessoais_pagas
        saldo = total_ins - total_outs

        print(f"   {'projetos_pessoais':20s}: €{float(projetos_pessoais):>12,.2f}")
        print(f"   {'premios':20s}: €{float(premios):>12,.2f}")
        print(f"   {'total_ins':20s}: €{float(total_ins):>12,.2f}")
        print(f"   {'fixas':20s}: €{float(despesas_fixas_metade):>12,.2f}")
        print(f"   {'pessoais_pagas':20s}: €{float(despesas_pessoais_pagas):>12,.2f}")
        print(f"   {'boletins_pagos':20s}: €{float(boletins_pagos):>12,.2f} (verificação)")
        print(f"   {'total_outs':20s}: €{float(total_outs):>12,.2f}")
        print(f"   {'SALDO':20s}: €{float(saldo):>12,.2f}")

    # 3. COMPARAÇÃO
    print("\n\n" + "=" * 100)
    print("📊 COMPARAÇÃO FINAL")
    print("=" * 100)

    print("\n👤 BRUNO:")
    print(f"   Excel  - Prémios: €{bruno_excel['premios']:,.2f} | Saldo: €{bruno_excel['saldo_sem_inv']:,.2f}")
    print(f"   DB     - Prémios: (calculado acima)")
    print(f"   Match: {'✅' if abs(bruno_excel['premios'] - float(premios)) < 0.01 else '❌'}")

    print("\n👤 RAFAEL:")
    print(f"   Excel  - Prémios: €{rafael_excel['premios']:,.2f} | Saldo: €{rafael_excel['saldo_sem_inv']:,.2f}")
    print(f"   DB     - Prémios: (calculado acima)")

except Exception as e:
    print(f"❌ Erro: {e}")
    import traceback
    traceback.print_exc()
finally:
    db.close()
