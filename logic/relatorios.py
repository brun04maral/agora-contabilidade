# -*- coding: utf-8 -*-
"""
Lógica de geração de relatórios
"""
from typing import Optional, Dict, Any, List
from sqlalchemy.orm import Session
from sqlalchemy import func, extract, or_
from datetime import date, datetime
from decimal import Decimal
from collections import defaultdict
from dateutil.relativedelta import relativedelta

from database.models import (
    Socio, Projeto, EstadoProjeto, TipoProjeto,
    Despesa, EstadoDespesa, TipoDespesa,
    Boletim, EstadoBoletim
)
from logic.saldos import SaldosCalculator


class RelatoriosManager:
    """
    Gestor de relatórios - geração e exportação
    """

    def __init__(self, db_session: Session):
        """
        Initialize manager

        Args:
            db_session: SQLAlchemy database session
        """
        self.db_session = db_session
        self.saldos_calculator = SaldosCalculator(db_session)

    def gerar_relatorio_saldos(
        self,
        socio: Optional[Socio] = None,
        data_inicio: Optional[date] = None,
        data_fim: Optional[date] = None,
        filtro_tipo_projeto: Optional[str] = None
    ) -> Dict[str, Any]:
        """
        Gera relatório de saldos pessoais

        Args:
            socio: Socio específico ou None para ambos
            data_inicio: Data de início do período (opcional)
            data_fim: Data de fim do período (opcional)
            filtro_tipo_projeto: Filtro de tipo ("todos", "empresa", "bruno", "rafael")

        Returns:
            Dicionário com dados do relatório
        """

        # Calcular saldos e buscar detalhes
        if socio == Socio.BRUNO:
            saldo_bruno = self.saldos_calculator.calcular_saldo_bruno()
            detalhes_bruno = self._get_detalhes_saldo_bruno(filtro_tipo_projeto)
            socios_data = [self._format_socio_data_detalhado("Bruno Amaral", saldo_bruno, detalhes_bruno, "#4CAF50")]
        elif socio == Socio.RAFAEL:
            saldo_rafael = self.saldos_calculator.calcular_saldo_rafael()
            detalhes_rafael = self._get_detalhes_saldo_rafael(filtro_tipo_projeto)
            socios_data = [self._format_socio_data_detalhado("Rafael Reigota", saldo_rafael, detalhes_rafael, "#efd578")]
        else:
            # Ambos
            saldo_bruno = self.saldos_calculator.calcular_saldo_bruno()
            detalhes_bruno = self._get_detalhes_saldo_bruno(filtro_tipo_projeto)
            saldo_rafael = self.saldos_calculator.calcular_saldo_rafael()
            detalhes_rafael = self._get_detalhes_saldo_rafael(filtro_tipo_projeto)
            socios_data = [
                self._format_socio_data_detalhado("Bruno Amaral", saldo_bruno, detalhes_bruno, "#4CAF50"),
                self._format_socio_data_detalhado("Rafael Reigota", saldo_rafael, detalhes_rafael, "#efd578")
            ]

        # Build report data
        periodo_str = self._format_periodo(data_inicio, data_fim)

        return {
            'tipo': 'saldos_pessoais',
            'titulo': 'Relatório de Saldos Pessoais',
            'periodo': periodo_str,
            'data_geracao': datetime.now().strftime('%d/%m/%Y %H:%M'),
            'data_inicio': data_inicio,
            'data_fim': data_fim,
            'socios': socios_data
        }

    def gerar_relatorio_financeiro_mensal(
        self,
        data_inicio: Optional[date] = None,
        data_fim: Optional[date] = None
    ) -> Dict[str, Any]:
        """
        Gera relatório financeiro mensal (faturação vs despesas)

        Args:
            data_inicio: Data de início do período
            data_fim: Data de fim do período

        Returns:
            Dicionário com dados do relatório
        """

        # Define período (último ano se não especificado)
        if not data_fim:
            data_fim = date.today()
        if not data_inicio:
            data_inicio = data_fim - relativedelta(years=1)

        # Buscar projetos recebidos no período
        projetos = self.db_session.query(Projeto).filter(
            Projeto.estado == EstadoProjeto.RECEBIDO,
            Projeto.data_faturacao.isnot(None),
            Projeto.data_faturacao >= data_inicio,
            Projeto.data_faturacao <= data_fim
        ).all()

        # Buscar despesas pagas no período
        despesas = self.db_session.query(Despesa).filter(
            Despesa.estado == EstadoDespesa.PAGO,
            Despesa.data_pagamento.isnot(None),
            Despesa.data_pagamento >= data_inicio,
            Despesa.data_pagamento <= data_fim
        ).all()

        # Agregar por mês
        meses_data = defaultdict(lambda: {
            'faturacao': Decimal('0'),
            'despesas': Decimal('0'),
            'resultado': Decimal('0')
        })

        # Agregar projetos
        for projeto in projetos:
            if projeto.data_faturacao:
                mes_key = projeto.data_faturacao.strftime('%Y-%m')
                meses_data[mes_key]['faturacao'] += projeto.valor_sem_iva

        # Agregar despesas
        for despesa in despesas:
            if despesa.data_pagamento:
                mes_key = despesa.data_pagamento.strftime('%Y-%m')
                meses_data[mes_key]['despesas'] += despesa.valor_com_iva

        # Calcular resultados
        for mes_key in meses_data:
            meses_data[mes_key]['resultado'] = (
                meses_data[mes_key]['faturacao'] -
                meses_data[mes_key]['despesas']
            )

        # Ordenar por mês e formatar
        meses_ordenados = sorted(meses_data.keys())
        meses_formatados = []

        total_faturacao = Decimal('0')
        total_despesas = Decimal('0')

        for mes_key in meses_ordenados:
            ano, mes = mes_key.split('-')
            mes_nome = self._get_month_name(int(mes))
            data = meses_data[mes_key]

            total_faturacao += data['faturacao']
            total_despesas += data['despesas']

            meses_formatados.append({
                'mes_key': mes_key,
                'mes_nome': mes_nome,
                'ano': ano,
                'faturacao': float(data['faturacao']),
                'faturacao_fmt': self._format_currency(float(data['faturacao'])),
                'despesas': float(data['despesas']),
                'despesas_fmt': self._format_currency(float(data['despesas'])),
                'resultado': float(data['resultado']),
                'resultado_fmt': self._format_currency(float(data['resultado'])),
                'cor_resultado': '#4CAF50' if data['resultado'] >= 0 else '#F44336'
            })

        periodo_str = self._format_periodo(data_inicio, data_fim)
        total_resultado = total_faturacao - total_despesas

        return {
            'tipo': 'financeiro_mensal',
            'titulo': 'Relatório Financeiro Mensal',
            'periodo': periodo_str,
            'data_geracao': datetime.now().strftime('%d/%m/%Y %H:%M'),
            'data_inicio': data_inicio,
            'data_fim': data_fim,
            'meses': meses_formatados,
            'totais': {
                'faturacao': float(total_faturacao),
                'faturacao_fmt': self._format_currency(float(total_faturacao)),
                'despesas': float(total_despesas),
                'despesas_fmt': self._format_currency(float(total_despesas)),
                'resultado': float(total_resultado),
                'resultado_fmt': self._format_currency(float(total_resultado)),
                'cor_resultado': '#4CAF50' if total_resultado >= 0 else '#F44336'
            }
        }

    def _get_month_name(self, mes: int) -> str:
        """Get month name in Portuguese"""
        meses = [
            'Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho',
            'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro'
        ]
        return meses[mes - 1]

    def gerar_relatorio_projetos(
        self,
        tipo: Optional['TipoProjeto'] = None,
        estado: Optional['EstadoProjeto'] = None,
        data_inicio: Optional[date] = None,
        data_fim: Optional[date] = None
    ) -> Dict[str, Any]:
        """
        Gera relatório de projetos

        Args:
            tipo: Filtrar por tipo de projeto (opcional)
            estado: Filtrar por estado (opcional)
            data_inicio: Data de início do período (opcional)
            data_fim: Data de fim do período (opcional)

        Returns:
            Dicionário com dados do relatório
        """
        from database.models import TipoProjeto, EstadoProjeto

        # Base query
        query = self.db_session.query(Projeto)

        # Apply filters
        if tipo:
            query = query.filter(Projeto.tipo == tipo)
        if estado:
            query = query.filter(Projeto.estado == estado)
        if data_inicio:
            query = query.filter(Projeto.data_inicio >= data_inicio)
        if data_fim:
            query = query.filter(Projeto.data_inicio <= data_fim)

        projetos = query.all()

        # Calculate statistics
        stats_por_tipo = {
            TipoProjeto.EMPRESA: {'count': 0, 'valor': Decimal('0'), 'premios': Decimal('0')},
            TipoProjeto.PESSOAL_BRUNO: {'count': 0, 'valor': Decimal('0'), 'premios': Decimal('0')},
            TipoProjeto.PESSOAL_RAFAEL: {'count': 0, 'valor': Decimal('0'), 'premios': Decimal('0')}
        }

        stats_por_estado = {
            EstadoProjeto.NAO_FATURADO: {'count': 0, 'valor': Decimal('0')},
            EstadoProjeto.FATURADO: {'count': 0, 'valor': Decimal('0')},
            EstadoProjeto.RECEBIDO: {'count': 0, 'valor': Decimal('0')},
            EstadoProjeto.ANULADO: {'count': 0, 'valor': Decimal('0')}
        }

        total_valor = Decimal('0')
        total_premios_bruno = Decimal('0')
        total_premios_rafael = Decimal('0')

        projetos_formatados = []

        for projeto in projetos:
            # Stats por tipo
            stats_por_tipo[projeto.tipo]['count'] += 1
            stats_por_tipo[projeto.tipo]['valor'] += projeto.valor_sem_iva
            if projeto.tipo == TipoProjeto.EMPRESA:
                stats_por_tipo[projeto.tipo]['premios'] += (projeto.premio_bruno + projeto.premio_rafael)

            # Stats por estado
            stats_por_estado[projeto.estado]['count'] += 1
            stats_por_estado[projeto.estado]['valor'] += projeto.valor_sem_iva

            # Totals
            total_valor += projeto.valor_sem_iva
            total_premios_bruno += projeto.premio_bruno
            total_premios_rafael += projeto.premio_rafael

            # Format project for table
            projetos_formatados.append({
                'numero': projeto.numero,
                'tipo': self._get_tipo_label(projeto.tipo),
                'cliente': projeto.cliente.nome if projeto.cliente else '-',
                'descricao': projeto.descricao[:40] + '...' if len(projeto.descricao) > 40 else projeto.descricao,
                'valor': float(projeto.valor_sem_iva),
                'valor_fmt': self._format_currency(float(projeto.valor_sem_iva)),
                'estado': self._get_estado_label(projeto.estado),
                'premio_bruno': float(projeto.premio_bruno) if projeto.premio_bruno else 0,
                'premio_bruno_fmt': self._format_currency(float(projeto.premio_bruno)) if projeto.premio_bruno else '-',
                'premio_rafael': float(projeto.premio_rafael) if projeto.premio_rafael else 0,
                'premio_rafael_fmt': self._format_currency(float(projeto.premio_rafael)) if projeto.premio_rafael else '-',
            })

        # Format statistics
        stats_tipo_fmt = []
        for tipo_proj, stats in stats_por_tipo.items():
            stats_tipo_fmt.append({
                'tipo': self._get_tipo_label(tipo_proj),
                'count': stats['count'],
                'valor': float(stats['valor']),
                'valor_fmt': self._format_currency(float(stats['valor'])),
                'premios': float(stats['premios']),
                'premios_fmt': self._format_currency(float(stats['premios'])) if stats['premios'] > 0 else '-'
            })

        stats_estado_fmt = []
        for estado_proj, stats in stats_por_estado.items():
            stats_estado_fmt.append({
                'estado': self._get_estado_label(estado_proj),
                'count': stats['count'],
                'valor': float(stats['valor']),
                'valor_fmt': self._format_currency(float(stats['valor']))
            })

        periodo_str = self._format_periodo(data_inicio, data_fim)

        # Determinar se deve mostrar prémios (só para Todos ou Empresa)
        mostrar_premios = tipo is None or tipo == TipoProjeto.EMPRESA

        return {
            'tipo': 'projetos',
            'titulo': 'Relatório de Projetos',
            'periodo': periodo_str,
            'data_geracao': datetime.now().strftime('%d/%m/%Y %H:%M'),
            'data_inicio': data_inicio,
            'data_fim': data_fim,
            'filtros': {
                'tipo': self._get_tipo_label(tipo) if tipo else 'Todos',
                'estado': self._get_estado_label(estado) if estado else 'Todos'
            },
            'mostrar_premios': mostrar_premios,
            'total_projetos': len(projetos),
            'total_valor': float(total_valor),
            'total_valor_fmt': self._format_currency(float(total_valor)),
            'total_premios_bruno': float(total_premios_bruno),
            'total_premios_bruno_fmt': self._format_currency(float(total_premios_bruno)),
            'total_premios_rafael': float(total_premios_rafael),
            'total_premios_rafael_fmt': self._format_currency(float(total_premios_rafael)),
            'stats_por_tipo': stats_tipo_fmt,
            'stats_por_estado': stats_estado_fmt,
            'projetos': projetos_formatados
        }

    def _get_tipo_label(self, tipo: 'TipoProjeto') -> str:
        """Get tipo label in Portuguese"""
        from database.models import TipoProjeto
        mapping = {
            TipoProjeto.EMPRESA: "Empresa",
            TipoProjeto.PESSOAL_BRUNO: "Pessoal Bruno",
            TipoProjeto.PESSOAL_RAFAEL: "Pessoal Rafael"
        }
        return mapping.get(tipo, str(tipo))

    def _get_estado_label(self, estado: 'EstadoProjeto') -> str:
        """Get estado label in Portuguese"""
        from database.models import EstadoProjeto
        mapping = {
            EstadoProjeto.NAO_FATURADO: "Não Faturado",
            EstadoProjeto.FATURADO: "Faturado",
            EstadoProjeto.RECEBIDO: "Recebido",
            EstadoProjeto.ANULADO: "Anulado"
        }
        return mapping.get(estado, str(estado))

    def _get_detalhes_saldo_bruno(self, filtro_tipo: Optional[str] = None) -> Dict[str, Any]:
        """Get detailed breakdown for Bruno's saldo with optional filter"""

        # Projetos pessoais recebidos (aplicar filtro)
        if filtro_tipo == "empresa":
            projetos_pessoais = []  # Não mostrar pessoais se filtro é empresa
        elif filtro_tipo == "rafael":
            projetos_pessoais = []  # Não mostrar Bruno se filtro é Rafael
        else:
            # "todos" ou "bruno" - mostrar projetos pessoais Bruno
            projetos_pessoais = self.db_session.query(Projeto).filter(
                Projeto.tipo == TipoProjeto.PESSOAL_BRUNO,
                Projeto.estado == EstadoProjeto.RECEBIDO
            ).all()

        # Prémios (projetos empresa onde Bruno tem prémio) - aplicar filtro
        if filtro_tipo in ["bruno", "rafael"]:
            projetos_premios = []  # Não mostrar prémios se filtro é só pessoais
        else:
            # "todos" ou "empresa" - mostrar prémios
            projetos_premios = self.db_session.query(Projeto).filter(
                Projeto.premio_bruno > 0
            ).all()

        # Despesas fixas pagas
        despesas_fixas = self.db_session.query(Despesa).filter(
            Despesa.tipo == TipoDespesa.FIXA_MENSAL,
            Despesa.estado == EstadoDespesa.PAGO
        ).all()

        # Boletins pagos
        boletins = self.db_session.query(Boletim).filter(
            Boletim.socio == Socio.BRUNO,
            Boletim.estado == EstadoBoletim.PAGO
        ).all()

        # Despesas pessoais pagas
        despesas_pessoais = self.db_session.query(Despesa).filter(
            Despesa.tipo == TipoDespesa.PESSOAL_BRUNO,
            Despesa.estado == EstadoDespesa.PAGO
        ).all()

        return {
            'projetos_pessoais': projetos_pessoais,
            'projetos_premios': projetos_premios,
            'despesas_fixas': despesas_fixas,
            'boletins': boletins,
            'despesas_pessoais': despesas_pessoais
        }

    def _get_detalhes_saldo_rafael(self, filtro_tipo: Optional[str] = None) -> Dict[str, Any]:
        """Get detailed breakdown for Rafael's saldo with optional filter"""

        # Projetos pessoais recebidos (aplicar filtro)
        if filtro_tipo == "empresa":
            projetos_pessoais = []  # Não mostrar pessoais se filtro é empresa
        elif filtro_tipo == "bruno":
            projetos_pessoais = []  # Não mostrar Rafael se filtro é Bruno
        else:
            # "todos" ou "rafael" - mostrar projetos pessoais Rafael
            projetos_pessoais = self.db_session.query(Projeto).filter(
                Projeto.tipo == TipoProjeto.PESSOAL_RAFAEL,
                Projeto.estado == EstadoProjeto.RECEBIDO
            ).all()

        # Prémios (projetos empresa onde Rafael tem prémio) - aplicar filtro
        if filtro_tipo in ["bruno", "rafael"]:
            projetos_premios = []  # Não mostrar prémios se filtro é só pessoais
        else:
            # "todos" ou "empresa" - mostrar prémios
            projetos_premios = self.db_session.query(Projeto).filter(
                Projeto.premio_rafael > 0
            ).all()

        # Despesas fixas pagas
        despesas_fixas = self.db_session.query(Despesa).filter(
            Despesa.tipo == TipoDespesa.FIXA_MENSAL,
            Despesa.estado == EstadoDespesa.PAGO
        ).all()

        # Boletins pagos
        boletins = self.db_session.query(Boletim).filter(
            Boletim.socio == Socio.RAFAEL,
            Boletim.estado == EstadoBoletim.PAGO
        ).all()

        # Despesas pessoais pagas
        despesas_pessoais = self.db_session.query(Despesa).filter(
            Despesa.tipo == TipoDespesa.PESSOAL_RAFAEL,
            Despesa.estado == EstadoDespesa.PAGO
        ).all()

        return {
            'projetos_pessoais': projetos_pessoais,
            'projetos_premios': projetos_premios,
            'despesas_fixas': despesas_fixas,
            'boletins': boletins,
            'despesas_pessoais': despesas_pessoais
        }

    def _format_socio_data_detalhado(self, nome: str, saldo_data: dict, detalhes: dict, cor: str) -> Dict[str, Any]:
        """Format socio data with detailed breakdowns"""

        ins = saldo_data['ins']
        outs = saldo_data['outs']

        # Format projetos pessoais list
        projetos_pessoais_list = []
        for proj in detalhes['projetos_pessoais']:
            projetos_pessoais_list.append({
                'numero': proj.numero,
                'cliente': proj.cliente.nome if proj.cliente else '-',
                'descricao': proj.descricao[:50] + '...' if len(proj.descricao) > 50 else proj.descricao,
                'valor': float(proj.valor_sem_iva),
                'valor_fmt': self._format_currency(float(proj.valor_sem_iva)),
                'data': proj.data_faturacao.strftime('%d/%m/%Y') if proj.data_faturacao else '-'
            })

        # Format prémios list
        premios_list = []
        for proj in detalhes['projetos_premios']:
            premio_bruno = proj.premio_bruno if 'Bruno' in nome else proj.premio_rafael
            if premio_bruno > 0:
                premios_list.append({
                    'numero': proj.numero,
                    'cliente': proj.cliente.nome if proj.cliente else '-',
                    'descricao': proj.descricao[:50] + '...' if len(proj.descricao) > 50 else proj.descricao,
                    'premio': float(premio_bruno),
                    'premio_fmt': self._format_currency(float(premio_bruno)),
                    'tipo': 'Cachet' if premio_bruno == proj.valor_sem_iva * Decimal('0.15') else 'Comissão'
                })

        # Format despesas fixas list
        despesas_fixas_list = []
        for desp in detalhes['despesas_fixas']:
            valor_metade = desp.valor_sem_iva / 2
            despesas_fixas_list.append({
                'numero': desp.numero,
                'fornecedor': desp.credor.nome if desp.credor else '-',
                'descricao': desp.descricao[:50] + '...' if len(desp.descricao) > 50 else desp.descricao,
                'valor_total': float(desp.valor_sem_iva),
                'valor_total_fmt': self._format_currency(float(desp.valor_sem_iva)),
                'valor_50': float(valor_metade),
                'valor_50_fmt': self._format_currency(float(valor_metade)),
                'data': desp.data_pagamento.strftime('%d/%m/%Y') if desp.data_pagamento else '-'
            })

        # Format boletins list
        boletins_list = []
        for bol in detalhes['boletins']:
            boletins_list.append({
                'numero': bol.numero,
                'descricao': bol.descricao[:50] + '...' if len(bol.descricao) > 50 else bol.descricao,
                'valor': float(bol.valor),
                'valor_fmt': self._format_currency(float(bol.valor)),
                'data_emissao': bol.data_emissao.strftime('%d/%m/%Y') if bol.data_emissao else '-',
                'data_pagamento': bol.data_pagamento.strftime('%d/%m/%Y') if bol.data_pagamento else '-'
            })

        # Format despesas pessoais list
        despesas_pessoais_list = []
        for desp in detalhes['despesas_pessoais']:
            despesas_pessoais_list.append({
                'numero': desp.numero,
                'fornecedor': desp.credor.nome if desp.credor else '-',
                'descricao': desp.descricao[:50] + '...' if len(desp.descricao) > 50 else desp.descricao,
                'valor': float(desp.valor_sem_iva),
                'valor_fmt': self._format_currency(float(desp.valor_sem_iva)),
                'data': desp.data_pagamento.strftime('%d/%m/%Y') if desp.data_pagamento else '-'
            })

        return {
            'nome': nome,
            'saldo': self._format_currency(saldo_data['saldo_total']),
            'saldo_valor': saldo_data['saldo_total'],
            'cor': cor,
            'ins': [
                {'label': 'Projetos Pessoais', 'valor': self._format_currency(ins['projetos_pessoais'])},
                {'label': 'Prémios', 'valor': self._format_currency(ins['premios'])}
            ],
            'total_ins': self._format_currency(ins['total']),
            'total_ins_valor': ins['total'],
            'outs': [
                {'label': 'Despesas Fixas (50%)', 'valor': self._format_currency(outs['despesas_fixas'])},
                {'label': 'Boletins Pagos', 'valor': self._format_currency(outs['boletins'])},
                {'label': 'Despesas Pessoais', 'valor': self._format_currency(outs['despesas_pessoais'])}
            ],
            'total_outs': self._format_currency(outs['total']),
            'total_outs_valor': outs['total'],
            # Detailed lists
            'projetos_pessoais_list': projetos_pessoais_list,
            'premios_list': premios_list,
            'despesas_fixas_list': despesas_fixas_list,
            'boletins_list': boletins_list,
            'despesas_pessoais_list': despesas_pessoais_list
        }

    def _format_socio_data(self, nome: str, saldo_data: dict, cor: str) -> Dict[str, Any]:
        """Format socio data for report"""

        ins = saldo_data['ins']
        outs = saldo_data['outs']

        return {
            'nome': nome,
            'saldo': self._format_currency(saldo_data['saldo_total']),
            'saldo_valor': saldo_data['saldo_total'],
            'cor': cor,
            'ins': [
                {'label': 'Projetos Pessoais', 'valor': self._format_currency(ins['projetos_pessoais'])},
                {'label': 'Prémios', 'valor': self._format_currency(ins['premios'])}
            ],
            'total_ins': self._format_currency(ins['total']),
            'total_ins_valor': ins['total'],
            'outs': [
                {'label': 'Despesas Fixas (50%)', 'valor': self._format_currency(outs['despesas_fixas'])},
                {'label': 'Boletins Pagos', 'valor': self._format_currency(outs['boletins'])},
                {'label': 'Despesas Pessoais', 'valor': self._format_currency(outs['despesas_pessoais'])}
            ],
            'total_outs': self._format_currency(outs['total']),
            'total_outs_valor': outs['total']
        }

    def _format_currency(self, valor: float) -> str:
        """Format value as currency"""
        return f"€{valor:,.2f}".replace(",", " ").replace(".", ",").replace(" ", ".")

    def _format_periodo(self, data_inicio: Optional[date], data_fim: Optional[date]) -> str:
        """Format period string"""
        if data_inicio and data_fim:
            return f"Período: {data_inicio.strftime('%d/%m/%Y')} a {data_fim.strftime('%d/%m/%Y')}"
        elif data_inicio:
            return f"A partir de: {data_inicio.strftime('%d/%m/%Y')}"
        elif data_fim:
            return f"Até: {data_fim.strftime('%d/%m/%Y')}"
        else:
            return "Período: Acumulado (todo o histórico)"

    def exportar_pdf(self, report_data: Dict[str, Any], filename: str):
        """
        Exporta relatório para PDF

        Args:
            report_data: Dados do relatório
            filename: Nome do arquivo
        """
        tipo = report_data.get('tipo')

        if tipo == 'saldos_pessoais':
            self._exportar_pdf_saldos(report_data, filename)
        elif tipo == 'financeiro_mensal':
            self._exportar_pdf_financeiro(report_data, filename)
        elif tipo == 'projetos':
            self._exportar_pdf_projetos(report_data, filename)
        else:
            raise ValueError(f"Tipo de relatório não suportado: {tipo}")

    def _criar_header_pdf(self, styles):
        """Create PDF header with logo and company name"""
        from reportlab.platypus import Paragraph, Spacer, Table, TableStyle, Image
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
        from reportlab.lib.styles import ParagraphStyle
        import os

        elements = []

        # Try to use PNG logo, fallback to text if not found
        logo_path = os.path.join(os.path.dirname(__file__), "..", "media", "AGORA media production@0.5x.png")

        if os.path.exists(logo_path):
            # Use PNG logo
            logo_img = Image(logo_path, width=8*cm, height=2*cm, kind='proportional')

            header_data = [[logo_img]]
            header_table = Table(header_data, colWidths=[18*cm])
            header_table.setStyle(TableStyle([
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('LEFTPADDING', (0, 0), (-1, -1), 0),
                ('RIGHTPADDING', (0, 0), (-1, -1), 0),
                ('TOPPADDING', (0, 0), (-1, -1), 0),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
            ]))
        else:
            # Fallback to text logo
            logo_style = ParagraphStyle(
                'Logo',
                parent=styles['Normal'],
                fontSize=32,
                textColor=colors.HexColor('#efd578'),
                fontName='Helvetica-Bold',
                alignment=TA_LEFT,
                leading=32
            )

            company_style = ParagraphStyle(
                'Company',
                parent=styles['Normal'],
                fontSize=18,
                fontName='Helvetica-Bold',
                alignment=TA_LEFT,
                leading=22
            )

            header_data = [[
                [
                    Paragraph('<font name="Helvetica-Bold" size="32" color="#efd578">a</font>', logo_style),
                    Paragraph('AGORA<br/><font size="12" color="#efd578">Media Production</font>', company_style)
                ]
            ]]

            header_table = Table(header_data, colWidths=[18*cm])
            header_table.setStyle(TableStyle([
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('LEFTPADDING', (0, 0), (-1, -1), 0),
                ('RIGHTPADDING', (0, 0), (-1, -1), 0),
                ('TOPPADDING', (0, 0), (-1, -1), 0),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ]))

        elements.append(header_table)
        elements.append(Spacer(1, 0.3*cm))

        # Separator line
        line_table = Table([['']], colWidths=[18*cm], rowHeights=[0.1*cm])
        line_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#efd578')),
            ('TOPPADDING', (0, 0), (-1, -1), 0),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
        ]))
        elements.append(line_table)
        elements.append(Spacer(1, 0.5*cm))

        return elements

    def _exportar_pdf_saldos(self, report_data: Dict[str, Any], filename: str):
        """Export Saldos report to PDF"""
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_LEFT

        # Create PDF
        doc = SimpleDocTemplate(filename, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()

        # Add header with logo
        elements.extend(self._criar_header_pdf(styles))

        # Custom styles
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#efd578'),
            alignment=TA_CENTER,
            spaceAfter=12
        )

        subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=styles['Normal'],
            fontSize=12,
            textColor=colors.grey,
            alignment=TA_CENTER,
            spaceAfter=20
        )

        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=16,
            textColor=colors.HexColor('#424242'),
            spaceAfter=10
        )

        # Title
        elements.append(Paragraph(report_data['titulo'], title_style))

        # Period and generation date
        if report_data['periodo']:
            elements.append(Paragraph(report_data['periodo'], subtitle_style))
        elements.append(Paragraph(f"Gerado em: {report_data['data_geracao']}", subtitle_style))
        elements.append(Spacer(1, 0.5*cm))

        # For each socio
        for socio_data in report_data['socios']:
            # Socio name
            elements.append(Paragraph(socio_data['nome'], heading_style))

            # Saldo
            saldo_data = [
                ['SALDO ATUAL', socio_data['saldo']]
            ]
            saldo_table = Table(saldo_data, colWidths=[8*cm, 6*cm])
            saldo_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#E3F2FD')),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#1976D2')),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 18),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
                ('TOPPADDING', (0, 0), (-1, -1), 12),
            ]))
            elements.append(saldo_table)
            elements.append(Spacer(1, 0.5*cm))

            # INs table
            ins_data = [
                ['RECEITAS (INs)', ''],
            ]
            for item in socio_data['ins']:
                ins_data.append([item['label'], item['valor']])
            ins_data.append(['TOTAL INs', socio_data['total_ins']])

            ins_table = Table(ins_data, colWidths=[10*cm, 5*cm])
            ins_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4CAF50')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#E8F5E9')),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ]))
            elements.append(ins_table)
            elements.append(Spacer(1, 0.3*cm))

            # OUTs table
            outs_data = [
                ['DESPESAS (OUTs)', ''],
            ]
            for item in socio_data['outs']:
                outs_data.append([item['label'], item['valor']])
            outs_data.append(['TOTAL OUTs', socio_data['total_outs']])

            outs_table = Table(outs_data, colWidths=[10*cm, 5*cm])
            outs_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#F44336')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#FFEBEE')),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ]))
            elements.append(outs_table)
            elements.append(Spacer(1, 0.5*cm))

            # Helper para criar estilos de tabela detalhada
            def create_detail_table_style():
                return TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#424242')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('FONTSIZE', (0, 1), (-1, -1), 8),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                    ('TOPPADDING', (0, 0), (-1, -1), 6),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')]),
                ])

            # Detailed section heading style
            detail_heading_style = ParagraphStyle(
                'DetailHeading',
                parent=styles['Heading3'],
                fontSize=12,
                textColor=colors.HexColor('#1976D2'),
                spaceAfter=5,
                spaceBefore=10
            )

            # Add detailed lists
            elements.append(Paragraph(f"Detalhes - {socio_data['nome']}", heading_style))
            elements.append(Spacer(1, 0.3*cm))

            # Projetos Pessoais
            if socio_data.get('projetos_pessoais_list'):
                elements.append(Paragraph(f"💼 Projetos Pessoais ({len(socio_data['projetos_pessoais_list'])} items)", detail_heading_style))
                proj_data = [['Nº', 'Cliente', 'Valor', 'Data']]
                for proj in socio_data['projetos_pessoais_list']:
                    proj_data.append([
                        proj['numero'],
                        proj['cliente'][:30],
                        proj['valor_fmt'],
                        proj['data']
                    ])
                proj_table = Table(proj_data, colWidths=[2*cm, 6*cm, 3*cm, 3*cm])
                proj_table.setStyle(create_detail_table_style())
                elements.append(proj_table)
                elements.append(Spacer(1, 0.3*cm))

            # Prémios
            if socio_data.get('premios_list'):
                elements.append(Paragraph(f"🏆 Prémios ({len(socio_data['premios_list'])} items)", detail_heading_style))
                premio_data = [['Nº', 'Cliente', 'Prémio', 'Tipo']]
                for premio in socio_data['premios_list']:
                    premio_data.append([
                        premio['numero'],
                        premio['cliente'][:30],
                        premio['premio_fmt'],
                        premio['tipo']
                    ])
                premio_table = Table(premio_data, colWidths=[2*cm, 6*cm, 3*cm, 3*cm])
                premio_table.setStyle(create_detail_table_style())
                elements.append(premio_table)
                elements.append(Spacer(1, 0.3*cm))

            # Despesas Fixas (showing all, not just first 10)
            if socio_data.get('despesas_fixas_list'):
                elements.append(Paragraph(f"🏢 Despesas Fixas - 50% ({len(socio_data['despesas_fixas_list'])} items)", detail_heading_style))
                desp_data = [['Nº', 'Fornecedor', 'Valor 50%', 'Data']]
                for desp in socio_data['despesas_fixas_list']:
                    desp_data.append([
                        desp['numero'],
                        desp['fornecedor'][:30],
                        desp['valor_50_fmt'],
                        desp['data']
                    ])
                desp_table = Table(desp_data, colWidths=[2*cm, 6*cm, 3*cm, 3*cm])
                desp_table.setStyle(create_detail_table_style())
                elements.append(desp_table)
                elements.append(Spacer(1, 0.3*cm))

            # Boletins
            if socio_data.get('boletins_list'):
                elements.append(Paragraph(f"📄 Boletins Pagos ({len(socio_data['boletins_list'])} items)", detail_heading_style))
                bol_data = [['Nº', 'Descrição', 'Valor', 'Data Pag.']]
                for bol in socio_data['boletins_list']:
                    bol_data.append([
                        bol['numero'],
                        bol['descricao'][:35],
                        bol['valor_fmt'],
                        bol['data_pagamento']
                    ])
                bol_table = Table(bol_data, colWidths=[2*cm, 6*cm, 3*cm, 3*cm])
                bol_table.setStyle(create_detail_table_style())
                elements.append(bol_table)
                elements.append(Spacer(1, 0.3*cm))

            # Despesas Pessoais
            if socio_data.get('despesas_pessoais_list'):
                elements.append(Paragraph(f"💳 Despesas Pessoais ({len(socio_data['despesas_pessoais_list'])} items)", detail_heading_style))
                desp_pes_data = [['Nº', 'Fornecedor', 'Valor', 'Data']]
                for desp in socio_data['despesas_pessoais_list']:
                    desp_pes_data.append([
                        desp['numero'],
                        desp['fornecedor'][:30],
                        desp['valor_fmt'],
                        desp['data']
                    ])
                desp_pes_table = Table(desp_pes_data, colWidths=[2*cm, 6*cm, 3*cm, 3*cm])
                desp_pes_table.setStyle(create_detail_table_style())
                elements.append(desp_pes_table)
                elements.append(Spacer(1, 0.3*cm))

            elements.append(Spacer(1, 1*cm))

        # Build PDF
        doc.build(elements)

    def _exportar_pdf_financeiro(self, report_data: Dict[str, Any], filename: str):
        """Export Financeiro Mensal report to PDF"""
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER

        doc = SimpleDocTemplate(filename, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()

        # Add header with logo
        elements.extend(self._criar_header_pdf(styles))

        # Styles
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#efd578'),
            alignment=TA_CENTER,
            spaceAfter=12
        )

        subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=styles['Normal'],
            fontSize=12,
            textColor=colors.grey,
            alignment=TA_CENTER,
            spaceAfter=20
        )

        # Title and metadata
        elements.append(Paragraph(report_data['titulo'], title_style))
        if report_data['periodo']:
            elements.append(Paragraph(report_data['periodo'], subtitle_style))
        elements.append(Paragraph(f"Gerado em: {report_data['data_geracao']}", subtitle_style))
        elements.append(Spacer(1, 0.5*cm))

        # Monthly data table
        table_data = [['Mês', 'Faturação', 'Despesas', 'Resultado']]

        for mes in report_data['meses']:
            table_data.append([
                f"{mes['mes_nome']} {mes['ano']}",
                mes['faturacao_fmt'],
                mes['despesas_fmt'],
                mes['resultado_fmt']
            ])

        # Totals row
        totais = report_data['totais']
        table_data.append([
            'TOTAL',
            totais['faturacao_fmt'],
            totais['despesas_fmt'],
            totais['resultado_fmt']
        ])

        table = Table(table_data, colWidths=[5*cm, 4*cm, 4*cm, 4*cm])

        # Base style
        table_style = [
            # Header
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#efd578')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TOPPADDING', (0, 0), (-1, 0), 12),

            # Data rows
            ('ALIGN', (0, 1), (0, -2), 'LEFT'),
            ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
            ('FONTSIZE', (0, 1), (-1, -2), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#F5F5F5')]),

            # Total row
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#E3F2FD')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, -1), (-1, -1), 11),
            ('TOPPADDING', (0, -1), (-1, -1), 12),
            ('BOTTOMPADDING', (0, -1), (-1, -1), 12),
        ]

        # Color resultado column based on values
        for i, mes in enumerate(report_data['meses'], start=1):
            color = colors.HexColor('#4CAF50') if mes['resultado'] >= 0 else colors.HexColor('#F44336')
            table_style.append(('TEXTCOLOR', (3, i), (3, i), color))

        # Total resultado color
        total_color = colors.HexColor(totais['cor_resultado'])
        table_style.append(('TEXTCOLOR', (3, -1), (3, -1), total_color))

        table.setStyle(TableStyle(table_style))
        elements.append(table)

        doc.build(elements)

    def exportar_excel(self, report_data: Dict[str, Any], filename: str):
        """
        Exporta relatório para Excel

        Args:
            report_data: Dados do relatório
            filename: Nome do arquivo
        """
        tipo = report_data.get('tipo')

        if tipo == 'saldos_pessoais':
            self._exportar_excel_saldos(report_data, filename)
        elif tipo == 'financeiro_mensal':
            self._exportar_excel_financeiro(report_data, filename)
        elif tipo == 'projetos':
            self._exportar_excel_projetos(report_data, filename)
        else:
            raise ValueError(f"Tipo de relatório não suportado: {tipo}")

    def _exportar_excel_saldos(self, report_data: Dict[str, Any], filename: str):
        """Export Saldos report to Excel"""
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Saldos Pessoais"

        # Set column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 18

        row = 1

        # Title
        ws.merge_cells(f'A{row}:B{row}')
        cell = ws[f'A{row}']
        cell.value = report_data['titulo']
        cell.font = Font(size=18, bold=True, color="2196F3")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        row += 1

        # Period
        if report_data['periodo']:
            ws.merge_cells(f'A{row}:B{row}')
            cell = ws[f'A{row}']
            cell.value = report_data['periodo']
            cell.font = Font(size=11, color="666666")
            cell.alignment = Alignment(horizontal='center')
            row += 1

        # Generation date
        ws.merge_cells(f'A{row}:B{row}')
        cell = ws[f'A{row}']
        cell.value = f"Gerado em: {report_data['data_geracao']}"
        cell.font = Font(size=10, color="999999")
        cell.alignment = Alignment(horizontal='center')
        row += 2

        # For each socio
        for socio_data in report_data['socios']:
            # Socio name
            ws.merge_cells(f'A{row}:B{row}')
            cell = ws[f'A{row}']
            cell.value = socio_data['nome']
            cell.font = Font(size=14, bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
            row += 1

            # Saldo
            ws.merge_cells(f'A{row}:B{row}')
            cell = ws[f'A{row}']
            cell.value = f"SALDO ATUAL: {socio_data['saldo']}"
            cell.font = Font(size=16, bold=True, color="1976D2")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color="BBDEFB", end_color="BBDEFB", fill_type="solid")
            ws.row_dimensions[row].height = 30
            row += 2

            # INs header
            ws[f'A{row}'].value = "RECEITAS (INs)"
            ws[f'A{row}'].font = Font(bold=True, color="FFFFFF")
            ws[f'A{row}'].fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
            ws[f'B{row}'].fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
            row += 1

            # INs items
            for item in socio_data['ins']:
                ws[f'A{row}'].value = item['label']
                ws[f'B{row}'].value = item['valor']
                ws[f'B{row}'].alignment = Alignment(horizontal='right')
                row += 1

            # INs total
            ws[f'A{row}'].value = "TOTAL INs"
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'].value = socio_data['total_ins']
            ws[f'B{row}'].font = Font(bold=True)
            ws[f'B{row}'].alignment = Alignment(horizontal='right')
            ws[f'A{row}'].fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
            ws[f'B{row}'].fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
            row += 2

            # OUTs header
            ws[f'A{row}'].value = "DESPESAS (OUTs)"
            ws[f'A{row}'].font = Font(bold=True, color="FFFFFF")
            ws[f'A{row}'].fill = PatternFill(start_color="F44336", end_color="F44336", fill_type="solid")
            ws[f'B{row}'].fill = PatternFill(start_color="F44336", end_color="F44336", fill_type="solid")
            row += 1

            # OUTs items
            for item in socio_data['outs']:
                ws[f'A{row}'].value = item['label']
                ws[f'B{row}'].value = item['valor']
                ws[f'B{row}'].alignment = Alignment(horizontal='right')
                row += 1

            # OUTs total
            ws[f'A{row}'].value = "TOTAL OUTs"
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'].value = socio_data['total_outs']
            ws[f'B{row}'].font = Font(bold=True)
            ws[f'B{row}'].alignment = Alignment(horizontal='right')
            ws[f'A{row}'].fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
            ws[f'B{row}'].fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
            row += 2

            # Detailed Lists
            # First, adjust column widths for details
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 35
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 15

            # Detailed section title
            ws.merge_cells(f'A{row}:D{row}')
            cell = ws[f'A{row}']
            cell.value = f"Detalhes - {socio_data['nome']}"
            cell.font = Font(size=13, bold=True, color="1976D2")
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
            row += 2

            # Helper to create detail headers
            def create_detail_header(row_num, title, cols='A:D'):
                ws.merge_cells(f'{cols[0]}{row_num}:{cols[-1]}{row_num}')
                cell = ws[f'{cols[0]}{row_num}']
                cell.value = title
                cell.font = Font(size=11, bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="424242", end_color="424242", fill_type="solid")
                cell.alignment = Alignment(horizontal='left')
                return row_num + 1

            # Projetos Pessoais
            if socio_data.get('projetos_pessoais_list'):
                row = create_detail_header(row, f"💼 Projetos Pessoais ({len(socio_data['projetos_pessoais_list'])} items)")

                # Column headers
                ws[f'A{row}'].value = 'Nº'
                ws[f'B{row}'].value = 'Cliente'
                ws[f'C{row}'].value = 'Valor'
                ws[f'D{row}'].value = 'Data'
                for col in ['A', 'B', 'C', 'D']:
                    ws[f'{col}{row}'].font = Font(bold=True)
                    ws[f'{col}{row}'].fill = PatternFill(start_color="BDBDBD", end_color="BDBDBD", fill_type="solid")
                row += 1

                # Data rows
                for proj in socio_data['projetos_pessoais_list']:
                    ws[f'A{row}'].value = proj['numero']
                    ws[f'B{row}'].value = proj['cliente']
                    ws[f'C{row}'].value = proj['valor_fmt']
                    ws[f'C{row}'].alignment = Alignment(horizontal='right')
                    ws[f'D{row}'].value = proj['data']
                    row += 1
                row += 1

            # Prémios
            if socio_data.get('premios_list'):
                row = create_detail_header(row, f"🏆 Prémios ({len(socio_data['premios_list'])} items)")

                # Column headers
                ws[f'A{row}'].value = 'Nº'
                ws[f'B{row}'].value = 'Cliente'
                ws[f'C{row}'].value = 'Prémio'
                ws[f'D{row}'].value = 'Tipo'
                for col in ['A', 'B', 'C', 'D']:
                    ws[f'{col}{row}'].font = Font(bold=True)
                    ws[f'{col}{row}'].fill = PatternFill(start_color="BDBDBD", end_color="BDBDBD", fill_type="solid")
                row += 1

                # Data rows
                for premio in socio_data['premios_list']:
                    ws[f'A{row}'].value = premio['numero']
                    ws[f'B{row}'].value = premio['cliente']
                    ws[f'C{row}'].value = premio['premio_fmt']
                    ws[f'C{row}'].alignment = Alignment(horizontal='right')
                    ws[f'D{row}'].value = premio['tipo']
                    row += 1
                row += 1

            # Despesas Fixas
            if socio_data.get('despesas_fixas_list'):
                row = create_detail_header(row, f"🏢 Despesas Fixas - 50% ({len(socio_data['despesas_fixas_list'])} items)")

                # Column headers
                ws[f'A{row}'].value = 'Nº'
                ws[f'B{row}'].value = 'Fornecedor'
                ws[f'C{row}'].value = 'Valor 50%'
                ws[f'D{row}'].value = 'Data'
                for col in ['A', 'B', 'C', 'D']:
                    ws[f'{col}{row}'].font = Font(bold=True)
                    ws[f'{col}{row}'].fill = PatternFill(start_color="BDBDBD", end_color="BDBDBD", fill_type="solid")
                row += 1

                # Data rows
                for desp in socio_data['despesas_fixas_list']:
                    ws[f'A{row}'].value = desp['numero']
                    ws[f'B{row}'].value = desp['fornecedor']
                    ws[f'C{row}'].value = desp['valor_50_fmt']
                    ws[f'C{row}'].alignment = Alignment(horizontal='right')
                    ws[f'D{row}'].value = desp['data']
                    row += 1
                row += 1

            # Boletins
            if socio_data.get('boletins_list'):
                row = create_detail_header(row, f"📄 Boletins Pagos ({len(socio_data['boletins_list'])} items)")

                # Column headers
                ws[f'A{row}'].value = 'Nº'
                ws[f'B{row}'].value = 'Descrição'
                ws[f'C{row}'].value = 'Valor'
                ws[f'D{row}'].value = 'Data Pag.'
                for col in ['A', 'B', 'C', 'D']:
                    ws[f'{col}{row}'].font = Font(bold=True)
                    ws[f'{col}{row}'].fill = PatternFill(start_color="BDBDBD", end_color="BDBDBD", fill_type="solid")
                row += 1

                # Data rows
                for bol in socio_data['boletins_list']:
                    ws[f'A{row}'].value = bol['numero']
                    ws[f'B{row}'].value = bol['descricao']
                    ws[f'C{row}'].value = bol['valor_fmt']
                    ws[f'C{row}'].alignment = Alignment(horizontal='right')
                    ws[f'D{row}'].value = bol['data_pagamento']
                    row += 1
                row += 1

            # Despesas Pessoais
            if socio_data.get('despesas_pessoais_list'):
                row = create_detail_header(row, f"💳 Despesas Pessoais ({len(socio_data['despesas_pessoais_list'])} items)")

                # Column headers
                ws[f'A{row}'].value = 'Nº'
                ws[f'B{row}'].value = 'Fornecedor'
                ws[f'C{row}'].value = 'Valor'
                ws[f'D{row}'].value = 'Data'
                for col in ['A', 'B', 'C', 'D']:
                    ws[f'{col}{row}'].font = Font(bold=True)
                    ws[f'{col}{row}'].fill = PatternFill(start_color="BDBDBD", end_color="BDBDBD", fill_type="solid")
                row += 1

                # Data rows
                for desp in socio_data['despesas_pessoais_list']:
                    ws[f'A{row}'].value = desp['numero']
                    ws[f'B{row}'].value = desp['fornecedor']
                    ws[f'C{row}'].value = desp['valor_fmt']
                    ws[f'C{row}'].alignment = Alignment(horizontal='right')
                    ws[f'D{row}'].value = desp['data']
                    row += 1
                row += 1

            row += 2

        # Save
        wb.save(filename)

    def _exportar_excel_financeiro(self, report_data: Dict[str, Any], filename: str):
        """Export Financeiro Mensal report to Excel"""
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Financeiro Mensal"

        # Column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15

        row = 1

        # Title
        ws.merge_cells(f'A{row}:D{row}')
        cell = ws[f'A{row}']
        cell.value = report_data['titulo']
        cell.font = Font(size=18, bold=True, color="2196F3")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        row += 1

        # Period
        if report_data['periodo']:
            ws.merge_cells(f'A{row}:D{row}')
            cell = ws[f'A{row}']
            cell.value = report_data['periodo']
            cell.font = Font(size=11, color="666666")
            cell.alignment = Alignment(horizontal='center')
            row += 1

        # Generation date
        ws.merge_cells(f'A{row}:D{row}')
        cell = ws[f'A{row}']
        cell.value = f"Gerado em: {report_data['data_geracao']}"
        cell.font = Font(size=10, color="999999")
        cell.alignment = Alignment(horizontal='center')
        row += 2

        # Header
        headers = ['Mês', 'Faturação', 'Despesas', 'Resultado']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        row += 1

        # Data rows
        for mes in report_data['meses']:
            ws[f'A{row}'].value = f"{mes['mes_nome']} {mes['ano']}"
            ws[f'B{row}'].value = mes['faturacao_fmt']
            ws[f'B{row}'].alignment = Alignment(horizontal='right')
            ws[f'C{row}'].value = mes['despesas_fmt']
            ws[f'C{row}'].alignment = Alignment(horizontal='right')
            ws[f'D{row}'].value = mes['resultado_fmt']
            ws[f'D{row}'].alignment = Alignment(horizontal='right')

            # Color resultado
            color = "4CAF50" if mes['resultado'] >= 0 else "F44336"
            ws[f'D{row}'].font = Font(color=color)

            # Alternate row colors
            if row % 2 == 0:
                for col in ['A', 'B', 'C', 'D']:
                    ws[f'{col}{row}'].fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

            row += 1

        # Totals row
        totais = report_data['totais']
        ws[f'A{row}'].value = "TOTAL"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].value = totais['faturacao_fmt']
        ws[f'B{row}'].font = Font(bold=True)
        ws[f'B{row}'].alignment = Alignment(horizontal='right')
        ws[f'C{row}'].value = totais['despesas_fmt']
        ws[f'C{row}'].font = Font(bold=True)
        ws[f'C{row}'].alignment = Alignment(horizontal='right')
        ws[f'D{row}'].value = totais['resultado_fmt']
        ws[f'D{row}'].font = Font(bold=True, color=totais['cor_resultado'].replace('#', ''))
        ws[f'D{row}'].alignment = Alignment(horizontal='right')

        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{row}'].fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")

        wb.save(filename)

    def _exportar_pdf_projetos(self, report_data: Dict[str, Any], filename: str):
        """Export Projetos report to PDF"""
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER

        doc = SimpleDocTemplate(filename, pagesize=landscape(A4))
        elements = []
        styles = getSampleStyleSheet()

        # Add header with logo (landscape width)
        elements.extend(self._criar_header_pdf(styles))

        # Styles
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#efd578'),
            alignment=TA_CENTER,
            spaceAfter=12
        )

        subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=styles['Normal'],
            fontSize=12,
            textColor=colors.grey,
            alignment=TA_CENTER,
            spaceAfter=20
        )

        # Title and metadata
        elements.append(Paragraph(report_data['titulo'], title_style))
        if report_data['periodo']:
            elements.append(Paragraph(report_data['periodo'], subtitle_style))
        elements.append(Paragraph(f"Gerado em: {report_data['data_geracao']}", subtitle_style))
        elements.append(Spacer(1, 0.3*cm))

        # Summary stats
        summary_data = [[
            f"Total: {report_data['total_projetos']} projetos",
            f"Valor Total: {report_data['total_valor_fmt']}",
            f"Prémios Bruno: {report_data['total_premios_bruno_fmt']}",
            f"Prémios Rafael: {report_data['total_premios_rafael_fmt']}"
        ]]

        summary_table = Table(summary_data, colWidths=[5*cm, 5*cm, 5*cm, 5*cm])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#E3F2FD')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#1976D2')),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
            ('TOPPADDING', (0, 0), (-1, -1), 10),
        ]))
        elements.append(summary_table)
        elements.append(Spacer(1, 0.5*cm))

        # Projects table - show ALL projects
        table_data = [['Nº', 'Tipo', 'Cliente', 'Valor', 'Prémio B', 'Prémio R', 'Estado']]

        for proj in report_data['projetos']:  # Show all projects
            table_data.append([
                proj['numero'],
                proj['tipo'],
                proj['cliente'][:15],
                proj['valor_fmt'],
                proj.get('premio_bruno_fmt', '€0,00'),
                proj.get('premio_rafael_fmt', '€0,00'),
                proj['estado']
            ])

        table = Table(table_data, colWidths=[2*cm, 2.5*cm, 4.5*cm, 2.5*cm, 2.5*cm, 2.5*cm, 2.5*cm])

        table.setStyle(TableStyle([
            # Header
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#9C27B0')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TOPPADDING', (0, 0), (-1, 0), 12),
            # Data rows
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')]),
        ]))
        elements.append(table)

        doc.build(elements)

    def _exportar_excel_projetos(self, report_data: Dict[str, Any], filename: str):
        """Export Projetos report to Excel"""
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Projetos"

        # Column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 18

        row = 1

        # Title
        ws.merge_cells(f'A{row}:H{row}')
        cell = ws[f'A{row}']
        cell.value = report_data['titulo']
        cell.font = Font(size=18, bold=True, color="2196F3")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        row += 1

        # Period
        if report_data['periodo']:
            ws.merge_cells(f'A{row}:H{row}')
            cell = ws[f'A{row}']
            cell.value = report_data['periodo']
            cell.font = Font(size=11, color="666666")
            cell.alignment = Alignment(horizontal='center')
            row += 1

        # Generation date
        ws.merge_cells(f'A{row}:H{row}')
        cell = ws[f'A{row}']
        cell.value = f"Gerado em: {report_data['data_geracao']}"
        cell.font = Font(size=10, color="999999")
        cell.alignment = Alignment(horizontal='center')
        row += 2

        # Summary
        ws.merge_cells(f'A{row}:H{row}')
        cell = ws[f'A{row}']
        cell.value = f"Total: {report_data['total_projetos']} projetos | Valor: {report_data['total_valor_fmt']} | Prémios: Bruno {report_data['total_premios_bruno_fmt']} | Rafael {report_data['total_premios_rafael_fmt']}"
        cell.font = Font(size=11, bold=True)
        cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
        row += 2

        # Header
        headers = ['Nº', 'Tipo', 'Cliente', 'Descrição', 'Valor', 'Prémio Bruno', 'Prémio Rafael', 'Estado']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="9C27B0", end_color="9C27B0", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        row += 1

        # Data rows
        for proj in report_data['projetos']:
            ws[f'A{row}'].value = proj['numero']
            ws[f'B{row}'].value = proj['tipo']
            ws[f'C{row}'].value = proj['cliente']
            ws[f'D{row}'].value = proj['descricao']
            ws[f'E{row}'].value = proj['valor_fmt']
            ws[f'E{row}'].alignment = Alignment(horizontal='right')
            ws[f'F{row}'].value = proj.get('premio_bruno_fmt', '€0,00')
            ws[f'F{row}'].alignment = Alignment(horizontal='right')
            ws[f'G{row}'].value = proj.get('premio_rafael_fmt', '€0,00')
            ws[f'G{row}'].alignment = Alignment(horizontal='right')
            ws[f'H{row}'].value = proj['estado']

            # Alternate row colors
            if row % 2 == 0:
                for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
                    ws[f'{col}{row}'].fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

            row += 1

        wb.save(filename)
