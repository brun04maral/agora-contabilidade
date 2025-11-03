# -*- coding: utf-8 -*-
"""
Calcular saldos com a lógica CORRETA do Excel
"""
from dotenv import load_dotenv
import os
from sqlalchemy import create_engine, func
from sqlalchemy.orm import sessionmaker
from database.models import Projeto, TipoProjeto, EstadoProjeto, Despesa, TipoDespesa, EstadoDespesa, Boletim, Socio, EstadoBoletim
from decimal import Decimal
import openpyxl

load_dotenv()

print("=" * 100)
print("CÁLCULO DE SALDOS - LÓGICA CORRETA")
print("=" * 100)

# 1. Valores do Excel
wb = openpyxl.load_workbook('CONTABILIDADE_FINAL_20251102.xlsx', data_only=True)
ws = wb['CAIXA']

bruno_excel_saldo = ws.cell(row=4, column=13).value  # Coluna M (saldo sem investimento)
rafael_excel_saldo = ws.cell(row=5, column=13).value

print("\n📊 SALDOS DO EXCEL:")
print(f"   Bruno: €{bruno_excel_saldo:,.2f}")
print(f"   Rafael: €{rafael_excel_saldo:,.2f}")

# 2. Calcular na DB
database_url = os.getenv("DATABASE_URL", "sqlite:///./agora_media.db")
engine = create_engine(database_url)
Session = sessionmaker(bind=engine)
db = Session()

try:
    print("\n" + "=" * 100)
    print("CÁLCULO NA BASE DE DADOS")
    print("=" * 100)

    for socio_nome, socio_enum in [("BRUNO", Socio.BRUNO), ("RAFAEL", Socio.RAFAEL)]:
        print(f"\n👤 {socio_nome}:")
        print("-" * 100)

        # 1. Projetos pessoais RECEBIDOS
        if socio_nome == "BRUNO":
            projetos_pessoais = db.query(func.sum(Projeto.valor_sem_iva)).filter(
                Projeto.tipo == TipoProjeto.PESSOAL_BRUNO,
                Projeto.estado == EstadoProjeto.RECEBIDO
            ).scalar() or Decimal(0)
        else:
            projetos_pessoais = db.query(func.sum(Projeto.valor_sem_iva)).filter(
                Projeto.tipo == TipoProjeto.PESSOAL_RAFAEL,
                Projeto.estado == EstadoProjeto.RECEBIDO
            ).scalar() or Decimal(0)

        # 2. TODOS os prémios (não filtrar por estado!)
        if socio_nome == "BRUNO":
            premios = db.query(func.sum(Projeto.premio_bruno)).filter(
                Projeto.premio_bruno > 0
            ).scalar() or Decimal(0)
        else:
            premios = db.query(func.sum(Projeto.premio_rafael)).filter(
                Projeto.premio_rafael > 0
            ).scalar() or Decimal(0)

        # 3. Despesas fixas mensais PAGAS ÷ 2
        despesas_fixas = db.query(func.sum(Despesa.valor_sem_iva)).filter(
            Despesa.tipo == TipoDespesa.FIXA_MENSAL,
            Despesa.estado == EstadoDespesa.PAGO
        ).scalar() or Decimal(0)
        despesas_fixas_metade = despesas_fixas / 2

        # 4. Boletins PAGOS (único OUT além de fixas!)
        boletins_pagos = db.query(func.sum(Boletim.valor)).filter(
            Boletim.socio == socio_enum,
            Boletim.estado == EstadoBoletim.PAGO
        ).scalar() or Decimal(0)

        # CÁLCULO
        total_ins = projetos_pessoais + premios
        total_outs = despesas_fixas_metade + boletins_pagos
        saldo = total_ins - total_outs

        print(f"  📈 INs:")
        print(f"     Projetos pessoais (recebidos): €{float(projetos_pessoais):>12,.2f}")
        print(f"     Prémios (TODOS):                €{float(premios):>12,.2f}")
        print(f"     ────────────────────────────────────────────────")
        print(f"     TOTAL INs:                      €{float(total_ins):>12,.2f}")

        print(f"\n  📉 OUTs:")
        print(f"     Despesas fixas (÷2):            €{float(despesas_fixas_metade):>12,.2f}")
        print(f"     Boletins PAGOS:                 €{float(boletins_pagos):>12,.2f}")
        print(f"     ────────────────────────────────────────────────")
        print(f"     TOTAL OUTs:                     €{float(total_outs):>12,.2f}")

        print(f"\n  💰 SALDO:                          €{float(saldo):>12,.2f}")

        # Comparar com Excel
        excel_saldo = bruno_excel_saldo if socio_nome == "BRUNO" else rafael_excel_saldo
        diferenca = float(saldo) - excel_saldo
        match = "✅" if abs(diferenca) < 1 else "❌"

        print(f"\n  📊 Comparação:")
        print(f"     Excel:                          €{excel_saldo:>12,.2f}")
        print(f"     DB:                             €{float(saldo):>12,.2f}")
        print(f"     Diferença:                      €{diferenca:>12,.2f} {match}")

    print("\n" + "=" * 100)

except Exception as e:
    print(f"❌ Erro: {e}")
    import traceback
    traceback.print_exc()
finally:
    db.close()
