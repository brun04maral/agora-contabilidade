#!/usr/bin/env python
"""
Cross-check script between Excel and Database
"""
import openpyxl
from decimal import Decimal, InvalidOperation
import django
import os

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'agora_web.settings')
django.setup()

from core.models import Projeto, Despesa, Boletim

def safe_decimal(value):
    """Convert value to Decimal safely"""
    if value is None:
        return Decimal('0')
    try:
        str_val = str(value).replace(',', '.').strip()
        if not str_val or str_val == '':
            return Decimal('0')
        return Decimal(str_val)
    except (InvalidOperation, ValueError):
        return Decimal('0')

print("=" * 80)
print("CROSS-CHECK: EXCEL vs DATABASE")
print("=" * 80)
print()

# Load Excel
wb = openpyxl.load_workbook('/app/CONTABILIDADE_FINAL_20251231.xlsx', data_only=True)

# ============================================================================
# 1. PROJETOS
# ============================================================================
print("📁 PROJETOS")
print("-" * 80)

ws_proj = wb['PROJETOS']

# Find header row
header_row = None
for row_idx in range(1, 20):
    cell_val = ws_proj.cell(row_idx, 1).value
    if cell_val and 'Nº PROJETO' in str(cell_val):
        header_row = row_idx
        break

if not header_row:
    print("❌ Header row not found in PROJETOS sheet")
else:
    print(f"✅ Header row found at row {header_row}")

    excel_projetos = {}
    for row_idx in range(header_row + 1, ws_proj.max_row + 1):
        numero = ws_proj.cell(row_idx, 1).value
        if not numero or not str(numero).startswith('#P'):
            continue

        valor_sem_iva = safe_decimal(ws_proj.cell(row_idx, 11).value)  # Column K

        excel_projetos[str(numero)] = {
            'valor_sem_iva': valor_sem_iva,
        }

    # Get DB projetos
    db_projetos = {}
    for p in Projeto.objects.all():
        db_projetos[p.numero] = {
            'valor_sem_iva': p.valor_sem_iva,
        }

    print(f"Excel: {len(excel_projetos)} projetos")
    print(f"DB:    {len(db_projetos)} projetos")
    print()

    # Check missing
    missing_in_db = set(excel_projetos.keys()) - set(db_projetos.keys())
    missing_in_excel = set(db_projetos.keys()) - set(excel_projetos.keys())

    if missing_in_db:
        print(f"⚠️  {len(missing_in_db)} projetos no Excel mas NÃO na DB:")
        for num in sorted(list(missing_in_db))[:10]:
            val = excel_projetos[num]['valor_sem_iva']
            print(f"   - {num} (€{val})")
        if len(missing_in_db) > 10:
            print(f"   ... and {len(missing_in_db) - 10} more")
        print()

    if missing_in_excel:
        print(f"⚠️  {len(missing_in_excel)} projetos na DB mas NÃO no Excel:")
        for num in sorted(list(missing_in_excel))[:10]:
            val = db_projetos[num]['valor_sem_iva']
            print(f"   - {num} (€{val})")
        if len(missing_in_excel) > 10:
            print(f"   ... and {len(missing_in_excel) - 10} more")
        print()

    # Check value differences
    value_diffs = []
    for num in set(excel_projetos.keys()) & set(db_projetos.keys()):
        excel_val = excel_projetos[num]['valor_sem_iva']
        db_val = db_projetos[num]['valor_sem_iva']
        diff = abs(excel_val - db_val)
        if diff > Decimal('0.01'):
            value_diffs.append({
                'numero': num,
                'excel': excel_val,
                'db': db_val,
                'diff': diff
            })

    if value_diffs:
        print(f"⚠️  {len(value_diffs)} projetos com valores diferentes:")
        for item in sorted(value_diffs, key=lambda x: x['diff'], reverse=True)[:10]:
            print(f"   - {item['numero']}: Excel=€{item['excel']}, DB=€{item['db']}, Diff=€{item['diff']}")
        if len(value_diffs) > 10:
            print(f"   ... and {len(value_diffs) - 10} more")
        print()

    if not missing_in_db and not missing_in_excel and not value_diffs:
        print("✅ Projetos: Excel e DB estão 100% sincronizados!")
        print()

print()

# ============================================================================
# 2. DESPESAS
# ============================================================================
print("💰 DESPESAS")
print("-" * 80)

ws_desp = wb['DESPESAS']

# Header is at row 5
excel_despesas = {}
for row_idx in range(6, ws_desp.max_row + 1):
    numero = ws_desp.cell(row_idx, 1).value
    if not numero or not str(numero).startswith('#D'):
        continue

    valor_sem_iva = safe_decimal(ws_desp.cell(row_idx, 10).value)  # Column J
    valor_com_iva = safe_decimal(ws_desp.cell(row_idx, 13).value)  # Column M

    excel_despesas[str(numero)] = {
        'valor_sem_iva': valor_sem_iva,
        'valor_com_iva': valor_com_iva,
    }

# Get DB despesas
db_despesas = {}
for d in Despesa.objects.all():
    db_despesas[d.numero] = {
        'valor_sem_iva': d.valor_sem_iva,
        'valor_com_iva': d.valor_com_iva,
    }

print(f"Excel: {len(excel_despesas)} despesas")
print(f"DB:    {len(db_despesas)} despesas")
print()

# Check missing
missing_in_db = set(excel_despesas.keys()) - set(db_despesas.keys())
missing_in_excel = set(db_despesas.keys()) - set(excel_despesas.keys())

if missing_in_db:
    print(f"⚠️  {len(missing_in_db)} despesas no Excel mas NÃO na DB:")
    for num in sorted(list(missing_in_db))[:10]:
        val = excel_despesas[num]['valor_sem_iva']
        print(f"   - {num} (€{val})")
    if len(missing_in_db) > 10:
        print(f"   ... and {len(missing_in_db) - 10} more")
    print()

if missing_in_excel:
    print(f"⚠️  {len(missing_in_excel)} despesas na DB mas NÃO no Excel:")
    for num in sorted(list(missing_in_excel))[:10]:
        val = db_despesas[num]['valor_sem_iva']
        print(f"   - {num} (€{val})")
    if len(missing_in_excel) > 10:
        print(f"   ... and {len(missing_in_excel) - 10} more")
    print()

# Check value differences
value_diffs = []
for num in set(excel_despesas.keys()) & set(db_despesas.keys()):
    excel_val = excel_despesas[num]['valor_sem_iva']
    db_val = db_despesas[num]['valor_sem_iva']
    diff = abs(excel_val - db_val)
    if diff > Decimal('0.01'):
        value_diffs.append({
            'numero': num,
            'excel': excel_val,
            'db': db_val,
            'diff': diff
        })

if value_diffs:
    print(f"⚠️  {len(value_diffs)} despesas com valores diferentes:")
    for item in sorted(value_diffs, key=lambda x: x['diff'], reverse=True)[:10]:
        print(f"   - {item['numero']}: Excel=€{item['excel']}, DB=€{item['db']}, Diff=€{item['diff']}")
    if len(value_diffs) > 10:
        print(f"   ... and {len(value_diffs) - 10} more")
    print()

if not missing_in_db and not missing_in_excel and not value_diffs:
    print("✅ Despesas: Excel e DB estão 100% sincronizados!")
    print()

print()

# ============================================================================
# 3. BOLETINS
# ============================================================================
print("📋 BOLETINS")
print("-" * 80)

ws_bol = wb['BOLETINS']

# Find header row
header_row = None
for row_idx in range(1, 20):
    cell_val = ws_bol.cell(row_idx, 1).value
    if cell_val and 'Nº BOLETIM' in str(cell_val):
        header_row = row_idx
        break

if not header_row:
    print("❌ Header row not found in BOLETINS sheet")
else:
    excel_boletins = {}
    for row_idx in range(header_row + 1, ws_bol.max_row + 1):
        numero = ws_bol.cell(row_idx, 1).value
        if not numero or not str(numero).startswith('#B'):
            continue

        valor_sem_iva = safe_decimal(ws_bol.cell(row_idx, 11).value)  # Column K
        valor_com_iva = safe_decimal(ws_bol.cell(row_idx, 13).value)  # Column M

        excel_boletins[str(numero)] = {
            'valor_sem_iva': valor_sem_iva,
            'valor_com_iva': valor_com_iva,
        }

    # Get DB boletins
    db_boletins = {}
    for b in Boletim.objects.all():
        db_boletins[b.numero] = {
            'valor_sem_iva': b.valor_sem_iva,
            'valor_com_iva': b.valor_com_iva,
        }

    print(f"Excel: {len(excel_boletins)} boletins")
    print(f"DB:    {len(db_boletins)} boletins")
    print()

    # Check missing
    missing_in_db = set(excel_boletins.keys()) - set(db_boletins.keys())
    missing_in_excel = set(db_boletins.keys()) - set(excel_boletins.keys())

    if missing_in_db:
        print(f"⚠️  {len(missing_in_db)} boletins no Excel mas NÃO na DB:")
        for num in sorted(list(missing_in_db))[:10]:
            val = excel_boletins[num]['valor_sem_iva']
            print(f"   - {num} (€{val})")
        if len(missing_in_db) > 10:
            print(f"   ... and {len(missing_in_db) - 10} more")
        print()

    if missing_in_excel:
        print(f"⚠️  {len(missing_in_excel)} boletins na DB mas NÃO no Excel:")
        for num in sorted(list(missing_in_excel))[:10]:
            val = db_boletins[num]['valor_sem_iva']
            print(f"   - {num} (€{val})")
        if len(missing_in_excel) > 10:
            print(f"   ... and {len(missing_in_excel) - 10} more")
        print()

    # Check value differences
    value_diffs = []
    for num in set(excel_boletins.keys()) & set(db_boletins.keys()):
        excel_val = excel_boletins[num]['valor_sem_iva']
        db_val = db_boletins[num]['valor_sem_iva']
        diff = abs(excel_val - db_val)
        if diff > Decimal('0.01'):
            value_diffs.append({
                'numero': num,
                'excel': excel_val,
                'db': db_val,
                'diff': diff
            })

    if value_diffs:
        print(f"⚠️  {len(value_diffs)} boletins com valores diferentes:")
        for item in sorted(value_diffs, key=lambda x: x['diff'], reverse=True)[:10]:
            print(f"   - {item['numero']}: Excel=€{item['excel']}, DB=€{item['db']}, Diff=€{item['diff']}")
        if len(value_diffs) > 10:
            print(f"   ... and {len(value_diffs) - 10} more")
        print()

    if not missing_in_db and not missing_in_excel and not value_diffs:
        print("✅ Boletins: Excel e DB estão 100% sincronizados!")
        print()

wb.close()

print()
print("=" * 80)
print("CROSS-CHECK COMPLETO")
print("=" * 80)
