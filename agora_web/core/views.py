"""
Views for Agora Contabilidade core app.
"""
from pathlib import Path
from django.shortcuts import render
from django.contrib.admin.views.decorators import staff_member_required
from django.http import HttpResponse
from datetime import datetime
import markdown
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


@staff_member_required
def changelog_view(request):
    """
    Display the CHANGELOG.md file with version history.
    Only accessible to staff members.
    """
    # Read CHANGELOG.md from /app/ (copied during Docker build)
    changelog_path = Path('/app/CHANGELOG.md')

    try:
        with open(changelog_path, 'r', encoding='utf-8') as f:
            changelog_content = f.read()

        # Convert Markdown to HTML
        changelog_html = markdown.markdown(
            changelog_content,
            extensions=['fenced_code', 'tables', 'nl2br']
        )
    except FileNotFoundError:
        changelog_html = f"<p>CHANGELOG.md not found at {changelog_path}</p>"

    context = {
        'changelog_html': changelog_html,
        'title': 'Changelog - Agora Contabilidade',
    }

    return render(request, 'admin/changelog.html', context)


@staff_member_required
def fiscal_iva_view(request):
    """
    Página dedicada IVA com navegação trimestral.
    """
    from core.utils.fiscal import FiscalCalculator
    from core.models import Despesa, Projeto
    from datetime import datetime

    calculator = FiscalCalculator()
    hoje = datetime.now().date()

    # Get parameters
    ano = int(request.GET.get('ano', hoje.year))
    trimestre = int(request.GET.get('trimestre', ((hoje.month - 1) // 3) + 1))

    # Get available years
    anos_despesas = Despesa.objects.dates('data', 'year', order='ASC').values_list('data__year', flat=True)
    anos_projetos = Projeto.objects.dates('data_inicio', 'year', order='ASC').values_list('data_inicio__year', flat=True)
    anos_disponiveis = sorted(set(list(anos_despesas) + list(anos_projetos)))

    # Calculate IVA for selected period
    iva = calculator.calcular_iva_trimestral(ano, trimestre)
    iva_breakdown = calculator.breakdown_iva_por_tags(ano, trimestre)

    context = {
        'ano_atual': ano,
        'trimestre_atual': trimestre,
        'anos_disponiveis': anos_disponiveis,
        'iva': iva,
        'iva_breakdown': iva_breakdown,
    }

    return render(request, 'admin/core/fiscal/iva.html', context)


@staff_member_required
def fiscal_irs_view(request):
    """
    Página dedicada IRS com navegação mensal.
    """
    from core.utils.fiscal import FiscalCalculator
    from core.models import Despesa, Projeto
    from datetime import datetime

    calculator = FiscalCalculator()
    hoje = datetime.now().date()

    # Get parameters
    ano = int(request.GET.get('ano', hoje.year))
    mes = int(request.GET.get('mes', hoje.month))

    # Get available years
    anos_despesas = Despesa.objects.dates('data', 'year', order='ASC').values_list('data__year', flat=True)
    anos_projetos = Projeto.objects.dates('data_inicio', 'year', order='ASC').values_list('data_inicio__year', flat=True)
    anos_disponiveis = sorted(set(list(anos_despesas) + list(anos_projetos)))

    # Calculate IRS for selected period
    irs = calculator.calcular_irs_mensal(ano, mes)

    # Month names
    meses = [
        {'num': 1, 'nome': 'Janeiro'},
        {'num': 2, 'nome': 'Fevereiro'},
        {'num': 3, 'nome': 'Março'},
        {'num': 4, 'nome': 'Abril'},
        {'num': 5, 'nome': 'Maio'},
        {'num': 6, 'nome': 'Junho'},
        {'num': 7, 'nome': 'Julho'},
        {'num': 8, 'nome': 'Agosto'},
        {'num': 9, 'nome': 'Setembro'},
        {'num': 10, 'nome': 'Outubro'},
        {'num': 11, 'nome': 'Novembro'},
        {'num': 12, 'nome': 'Dezembro'},
    ]

    context = {
        'ano_atual': ano,
        'mes_atual': mes,
        'anos_disponiveis': anos_disponiveis,
        'meses': meses,
        'irs': irs,
    }

    return render(request, 'admin/core/fiscal/irs.html', context)


@staff_member_required
def fiscal_irc_view(request):
    """
    Página dedicada IRC com navegação anual.
    """
    from core.utils.fiscal import FiscalCalculator
    from core.models import Despesa, Projeto
    from datetime import datetime

    calculator = FiscalCalculator()
    hoje = datetime.now().date()

    # Get parameters
    ano = int(request.GET.get('ano', hoje.year))

    # Get available years
    anos_despesas = Despesa.objects.dates('data', 'year', order='ASC').values_list('data__year', flat=True)
    anos_projetos = Projeto.objects.dates('data_inicio', 'year', order='ASC').values_list('data_inicio__year', flat=True)
    anos_disponiveis = sorted(set(list(anos_despesas) + list(anos_projetos)))

    # Calculate IRC for selected year
    irc = calculator.estimar_irc_anual(ano)
    irc_breakdown = calculator.breakdown_irc_por_tags(ano)

    context = {
        'ano_atual': ano,
        'anos_disponiveis': anos_disponiveis,
        'irc': irc,
        'irc_breakdown': irc_breakdown,
    }

    return render(request, 'admin/core/fiscal/irc.html', context)


@staff_member_required
def export_fiscal_excel(request):
    """
    Export fiscal breakdown data to Excel.
    Generates comprehensive fiscal reports with IVA and IRC breakdowns.
    """
    from core.utils.fiscal import FiscalCalculator
    from decimal import Decimal

    # Get parameters from request
    ano = int(request.GET.get('ano', datetime.now().year))
    trimestre = int(request.GET.get('trimestre', ((datetime.now().month - 1) // 3) + 1))
    report_type = request.GET.get('type', 'iva')  # 'iva' or 'irc'

    # Initialize calculator
    calculator = FiscalCalculator()

    # Create workbook
    wb = Workbook()
    ws = wb.active

    # Define styles
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    alert_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    green_font = Font(color="10B981", bold=True)
    red_font = Font(color="EF4444", bold=True)
    yellow_font = Font(color="F59E0B", bold=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    if report_type == 'iva':
        # IVA Report
        ws.title = f"IVA_Q{trimestre}_{ano}"

        # Title
        ws['A1'] = f"Breakdown IVA - Q{trimestre}/{ano}"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:E1')

        # Get IVA data
        iva_data = calculator.calcular_iva_trimestral(ano, trimestre)
        breakdown = calculator.breakdown_iva_por_tags(ano, trimestre)

        # Summary section
        ws['A3'] = "RESUMO GERAL"
        ws['A3'].font = Font(bold=True, size=12)
        ws['A4'] = "Total IVA Dedutível:"
        ws['B4'] = float(iva_data['iva_dedutivel']['total'])
        ws['B4'].number_format = '€#,##0.00'
        ws['A5'] = "Total IVA Liquidado:"
        ws['B5'] = float(iva_data['iva_liquidado']['total'])
        ws['B5'].number_format = '€#,##0.00'
        ws['A6'] = "IVA a Pagar:"
        ws['B6'] = float(iva_data['iva_a_pagar'])
        ws['B6'].number_format = '€#,##0.00'
        ws['B6'].font = Font(bold=True)

        # Alert for untagged expenses
        if iva_data['iva_dedutivel'].get('despesas_sem_tag', 0) > 0:
            row = 8
            ws[f'A{row}'] = f"⚠️ ATENÇÃO: {iva_data['iva_dedutivel']['despesas_sem_tag']} despesa(s) sem tag IVA (assumido 100% dedutível)"
            ws[f'A{row}'].fill = alert_fill
            ws[f'A{row}'].font = Font(bold=True)
            ws.merge_cells(f'A{row}:E{row}')

        # Breakdown table header
        header_row = 10
        ws[f'A{header_row}'] = "Categoria Fiscal"
        ws[f'B{header_row}'] = "Código"
        ws[f'C{header_row}'] = "N.º Despesas"
        ws[f'D{header_row}'] = "IVA Bruto"
        ws[f'E{header_row}'] = "% Dedutível"
        ws[f'F{header_row}'] = "IVA Dedutível"

        # Apply header style
        for col in range(1, 7):
            cell = ws.cell(row=header_row, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        # Breakdown data
        row = header_row + 1
        for tag_codigo, data in sorted(breakdown.items()):
            ws[f'A{row}'] = data['nome']
            ws[f'B{row}'] = tag_codigo
            ws[f'C{row}'] = data['count']
            ws[f'D{row}'] = float(data['iva_bruto'])
            ws[f'D{row}'].number_format = '€#,##0.00'
            ws[f'E{row}'] = float(data['percentagem'])
            ws[f'E{row}'].number_format = '0"%"'
            ws[f'F{row}'] = float(data['iva_dedutivel'])
            ws[f'F{row}'].number_format = '€#,##0.00'

            # Color code percentages
            percentagem = float(data['percentagem'])
            if percentagem == 100:
                ws[f'E{row}'].font = green_font
                ws[f'F{row}'].font = green_font
            elif percentagem == 0:
                ws[f'E{row}'].font = red_font
                ws[f'F{row}'].font = red_font
            else:
                ws[f'E{row}'].font = yellow_font
                ws[f'F{row}'].font = yellow_font

            # Apply borders
            for col in range(1, 7):
                ws.cell(row=row, column=col).border = border

            row += 1

    else:
        # IRC Report
        ws.title = f"IRC_{ano}"

        # Title
        ws['A1'] = f"Breakdown IRC - Ano {ano}"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:E1')

        # Get IRC data
        irc_data = calculator.estimar_irc_anual(ano)
        breakdown = calculator.breakdown_irc_por_tags(ano)

        # Summary section
        ws['A3'] = "RESUMO GERAL"
        ws['A3'].font = Font(bold=True, size=12)
        ws['A4'] = "Total Receitas:"
        ws['B4'] = float(irc_data['receitas_total'])
        ws['B4'].number_format = '€#,##0.00'
        ws['A5'] = "Total Despesas (dedutíveis):"
        ws['B5'] = float(irc_data['despesas_total'])
        ws['B5'].number_format = '€#,##0.00'
        ws['A6'] = "Matéria Coletável:"
        ws['B6'] = float(irc_data['lucro_tributavel'])
        ws['B6'].number_format = '€#,##0.00'
        ws['B6'].font = Font(bold=True)
        ws['A7'] = "IRC Estimado (16%/20%):"
        ws['B7'] = float(irc_data['irc_total'])
        ws['B7'].number_format = '€#,##0.00'
        ws['B7'].font = Font(bold=True)

        # Alert for untagged expenses
        if irc_data.get('despesas_sem_tag_irc', 0) > 0:
            row = 9
            ws[f'A{row}'] = f"⚠️ ATENÇÃO: {irc_data['despesas_sem_tag_irc']} despesa(s) sem tag IRC (assumido 100% dedutível)"
            ws[f'A{row}'].fill = alert_fill
            ws[f'A{row}'].font = Font(bold=True)
            ws.merge_cells(f'A{row}:E{row}')

        # Breakdown table header
        header_row = 11
        ws[f'A{header_row}'] = "Categoria Fiscal"
        ws[f'B{header_row}'] = "Código"
        ws[f'C{header_row}'] = "N.º Despesas"
        ws[f'D{header_row}'] = "Valor Bruto"
        ws[f'E{header_row}'] = "% Dedutível"
        ws[f'F{header_row}'] = "Valor Dedutível"

        # Apply header style
        for col in range(1, 7):
            cell = ws.cell(row=header_row, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        # Breakdown data
        row = header_row + 1
        for tag_codigo, data in sorted(breakdown.items()):
            ws[f'A{row}'] = data['nome']
            ws[f'B{row}'] = tag_codigo
            ws[f'C{row}'] = data['count']
            ws[f'D{row}'] = float(data['valor_bruto'])
            ws[f'D{row}'].number_format = '€#,##0.00'
            ws[f'E{row}'] = float(data['percentagem'])
            ws[f'E{row}'].number_format = '0"%"'
            ws[f'F{row}'] = float(data['valor_dedutivel'])
            ws[f'F{row}'].number_format = '€#,##0.00'

            # Color code percentages
            percentagem = float(data['percentagem'])
            if percentagem == 100:
                ws[f'E{row}'].font = green_font
                ws[f'F{row}'].font = green_font
            elif percentagem == 0:
                ws[f'E{row}'].font = red_font
                ws[f'F{row}'].font = red_font
            else:
                ws[f'E{row}'].font = yellow_font
                ws[f'F{row}'].font = yellow_font

            # Apply borders
            for col in range(1, 7):
                ws.cell(row=row, column=col).border = border

            row += 1

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"Fiscal_{report_type.upper()}_Q{trimestre}_{ano}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    # Save workbook to response
    wb.save(response)

    return response
