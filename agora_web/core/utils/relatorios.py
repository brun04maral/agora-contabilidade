# -*- coding: utf-8 -*-
"""
Sistema de Geração de Relatórios para Agora Contabilidade

Suporta exportação de projetos, despesas, boletins e orçamentos
em formatos PDF e Excel com templates personalizáveis.

Autor: Agora Contabilidade
Data: 2026-01-12
"""
from io import BytesIO
from datetime import date, datetime
from decimal import Decimal
from typing import List, Dict, Any, Optional

from django.http import HttpResponse
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
from django.conf import settings
import os

import xlsxwriter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


class RelatorioBase:
    """
    Classe base para geração de relatórios
    """

    # Informações da empresa
    EMPRESA_NOME = "Amaral & Reigota - Produção Audiovisual, Lda"
    EMPRESA_MARCA = "Agora Media Production"
    EMPRESA_NIPC = "518 351 190"
    EMPRESA_MORADA = "Portugal"  # Atualizar com morada real se necessário

    def __init__(self):
        self.today = date.today()

    def _format_currency(self, value) -> str:
        """Formata valor como moeda EUR"""
        if value is None:
            return "€0,00"
        return f"€{value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

    def _format_date(self, date_value) -> str:
        """Formata data para formato PT"""
        if date_value is None:
            return "-"
        if isinstance(date_value, str):
            return date_value
        return date_value.strftime("%d/%m/%Y")

    def _format_number(self, value) -> str:
        """Formata número com separadores PT"""
        if value is None:
            return "0"
        return f"{value:,.0f}".replace(",", ".")


class RelatorioProjetos(RelatorioBase):
    """
    Gerador de relatórios para Projetos
    """

    def gerar_pdf(self, projetos, filtros: Optional[Dict[str, Any]] = None) -> HttpResponse:
        """
        Gera relatório PDF de projetos

        Args:
            projetos: QuerySet de projetos
            filtros: Dicionário com filtros aplicados (opcional)

        Returns:
            HttpResponse com PDF
        """
        buffer = BytesIO()

        # Configurar documento PDF em landscape para caber mais colunas
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            rightMargin=1*cm,
            leftMargin=1*cm,
            topMargin=1.5*cm,
            bottomMargin=1.5*cm
        )

        # Container para elementos do PDF
        elements = []

        # Estilos
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#1976d2'),
            spaceAfter=12,
            alignment=TA_CENTER
        )

        subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=styles['Normal'],
            fontSize=9,
            textColor=colors.grey,
            spaceAfter=6,
            alignment=TA_CENTER
        )

        # Logo (se existir)
        logo_path = os.path.join(settings.BASE_DIR.parent, 'media', 'logos', 'logo_sidebar.png')
        if os.path.exists(logo_path):
            try:
                logo = Image(logo_path, width=3*cm, height=3*cm)
                elements.append(logo)
                elements.append(Spacer(1, 0.2*cm))
            except:
                pass  # Se falhar, continua sem logo

        # Cabeçalho
        elements.append(Paragraph(self.EMPRESA_MARCA, title_style))
        elements.append(Paragraph(f"{self.EMPRESA_NOME} | NIPC: {self.EMPRESA_NIPC}", subtitle_style))
        elements.append(Spacer(1, 0.3*cm))

        # Título do relatório
        titulo = "Relatório de Projetos"
        if filtros:
            if filtros.get('tipo_relatorio'):
                titulo = f"Relatório de Projetos - {filtros['tipo_relatorio']}"

        elements.append(Paragraph(titulo, title_style))

        # Informações do relatório
        info_text = f"Gerado em: {self._format_date(self.today)}"
        if filtros:
            if filtros.get('data_inicio') and filtros.get('data_fim'):
                info_text += f" | Período: {self._format_date(filtros['data_inicio'])} - {self._format_date(filtros['data_fim'])}"
            if filtros.get('socio'):
                info_text += f" | Sócio: {filtros['socio']}"
            if filtros.get('estado'):
                info_text += f" | Estado: {filtros['estado']}"

        elements.append(Paragraph(info_text, subtitle_style))
        elements.append(Spacer(1, 0.5*cm))

        # Preparar dados da tabela
        data = [
            ['#', 'Tipo', 'Sócio', 'Cliente', 'Descrição', 'Valor s/ IVA', 'Prémio BA', 'Prémio RR', 'Estado', 'Data Fat.']
        ]

        total_valor = Decimal('0.00')
        total_premio_ba = Decimal('0.00')
        total_premio_rr = Decimal('0.00')

        for projeto in projetos:
            # Truncar cliente para evitar sobreposição
            cliente_nome = str(projeto.cliente) if projeto.cliente else '-'
            if len(cliente_nome) > 20:
                cliente_nome = cliente_nome[:17] + '...'

            data.append([
                projeto.numero,
                projeto.get_tipo_display(),
                str(projeto.socio) if projeto.socio else '-',
                cliente_nome,
                projeto.descricao[:40] + '...' if len(projeto.descricao) > 40 else projeto.descricao,
                self._format_currency(projeto.valor_sem_iva),
                self._format_currency(projeto.premio_bruno) if projeto.premio_bruno else '-',
                self._format_currency(projeto.premio_rafael) if projeto.premio_rafael else '-',
                projeto.get_estado_display(),
                self._format_date(projeto.data_faturacao)
            ])

            total_valor += projeto.valor_sem_iva or Decimal('0.00')
            total_premio_ba += projeto.premio_bruno or Decimal('0.00')
            total_premio_rr += projeto.premio_rafael or Decimal('0.00')

        # Linha de totais
        data.append([
            '', '', '', '', 'TOTAL:',
            self._format_currency(total_valor),
            self._format_currency(total_premio_ba),
            self._format_currency(total_premio_rr),
            '', ''
        ])

        # Criar tabela
        table = Table(data, colWidths=[
            2*cm,   # #
            2*cm,   # Tipo
            1.5*cm, # Sócio
            3*cm,   # Cliente
            6*cm,   # Descrição
            2.5*cm, # Valor
            2.2*cm, # Prémio BA
            2.2*cm, # Prémio RR
            2*cm,   # Estado
            2.2*cm  # Data
        ])

        # Estilo da tabela
        table.setStyle(TableStyle([
            # Cabeçalho
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1976d2')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),

            # Dados
            ('FONTNAME', (0, 1), (-1, -2), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -2), 8),
            ('ALIGN', (5, 1), (7, -1), 'RIGHT'),  # Valores alinhados à direita
            ('ALIGN', (0, 1), (0, -1), 'CENTER'),  # Número centralizado

            # Linha de totais
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#e3f2fd')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, -1), (-1, -1), 9),

            # Bordas
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),

            # Cores alternadas nas linhas
            *[('BACKGROUND', (0, i), (-1, i), colors.HexColor('#f5f5f5'))
              for i in range(2, len(data)-1, 2)]
        ]))

        elements.append(table)

        # Resumo
        elements.append(Spacer(1, 0.5*cm))
        resumo_text = f"Total de projetos: {len(projetos)}"
        elements.append(Paragraph(resumo_text, subtitle_style))

        # Gerar PDF
        doc.build(elements)

        # Preparar response
        buffer.seek(0)
        filename = self._gerar_nome_arquivo(filtros, 'pdf')

        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        return response

    def gerar_excel(self, projetos, filtros: Optional[Dict[str, Any]] = None) -> HttpResponse:
        """
        Gera relatório Excel de projetos

        Args:
            projetos: QuerySet de projetos
            filtros: Dicionário com filtros aplicados (opcional)

        Returns:
            HttpResponse com Excel
        """
        buffer = BytesIO()

        # Criar workbook
        workbook = xlsxwriter.Workbook(buffer, {'in_memory': True})
        worksheet = workbook.add_worksheet('Projetos')

        # Formatos
        header_format = workbook.add_format({
            'bold': True,
            'bg_color': '#1976d2',
            'font_color': 'white',
            'align': 'center',
            'valign': 'vcenter',
            'border': 1
        })

        currency_format = workbook.add_format({
            'num_format': '€#,##0.00',
            'align': 'right'
        })

        date_format = workbook.add_format({
            'num_format': 'dd/mm/yyyy',
            'align': 'center'
        })

        total_format = workbook.add_format({
            'bold': True,
            'bg_color': '#e3f2fd',
            'num_format': '€#,##0.00',
            'align': 'right',
            'border': 1
        })

        text_format = workbook.add_format({
            'align': 'left',
            'valign': 'vcenter'
        })

        center_format = workbook.add_format({
            'align': 'center',
            'valign': 'vcenter'
        })

        # Informações do cabeçalho
        worksheet.write(0, 0, self.EMPRESA_MARCA, workbook.add_format({'bold': True, 'font_size': 14}))
        worksheet.write(1, 0, f"{self.EMPRESA_NOME} | NIPC: {self.EMPRESA_NIPC}", workbook.add_format({'font_size': 10}))

        titulo = "Relatório de Projetos"
        if filtros and filtros.get('tipo_relatorio'):
            titulo = f"Relatório de Projetos - {filtros['tipo_relatorio']}"
        worksheet.write(2, 0, titulo, workbook.add_format({'bold': True, 'font_size': 12}))

        info_text = f"Gerado em: {self.today.strftime('%d/%m/%Y')}"
        worksheet.write(3, 0, info_text, workbook.add_format({'font_size': 10}))

        # Cabeçalhos da tabela
        row = 5
        headers = ['Número', 'Tipo', 'Sócio', 'Cliente', 'Descrição', 'Valor s/ IVA',
                   'Prémio Bruno', 'Prémio Rafael', 'Estado', 'Data Faturação', 'Data Início', 'Data Fim']

        for col, header in enumerate(headers):
            worksheet.write(row, col, header, header_format)

        # Dados
        row += 1
        total_valor = Decimal('0.00')
        total_premio_ba = Decimal('0.00')
        total_premio_rr = Decimal('0.00')

        for projeto in projetos:
            worksheet.write(row, 0, projeto.numero, center_format)
            worksheet.write(row, 1, projeto.get_tipo_display(), center_format)
            worksheet.write(row, 2, str(projeto.socio) if projeto.socio else '-', center_format)
            worksheet.write(row, 3, str(projeto.cliente) if projeto.cliente else '-', text_format)
            worksheet.write(row, 4, projeto.descricao, text_format)
            worksheet.write(row, 5, float(projeto.valor_sem_iva or 0), currency_format)
            worksheet.write(row, 6, float(projeto.premio_bruno or 0), currency_format)
            worksheet.write(row, 7, float(projeto.premio_rafael or 0), currency_format)
            worksheet.write(row, 8, projeto.get_estado_display(), center_format)

            if projeto.data_faturacao:
                worksheet.write_datetime(row, 9, projeto.data_faturacao, date_format)
            else:
                worksheet.write(row, 9, '-', center_format)

            if projeto.data_inicio:
                worksheet.write_datetime(row, 10, projeto.data_inicio, date_format)
            else:
                worksheet.write(row, 10, '-', center_format)

            if projeto.data_fim:
                worksheet.write_datetime(row, 11, projeto.data_fim, date_format)
            else:
                worksheet.write(row, 11, '-', center_format)

            total_valor += projeto.valor_sem_iva or Decimal('0.00')
            total_premio_ba += projeto.premio_bruno or Decimal('0.00')
            total_premio_rr += projeto.premio_rafael or Decimal('0.00')

            row += 1

        # Linha de totais
        worksheet.write(row, 4, 'TOTAL:', total_format)
        worksheet.write(row, 5, float(total_valor), total_format)
        worksheet.write(row, 6, float(total_premio_ba), total_format)
        worksheet.write(row, 7, float(total_premio_rr), total_format)

        # Ajustar larguras das colunas
        worksheet.set_column(0, 0, 12)  # Número
        worksheet.set_column(1, 1, 12)  # Tipo
        worksheet.set_column(2, 2, 8)   # Sócio
        worksheet.set_column(3, 3, 20)  # Cliente
        worksheet.set_column(4, 4, 40)  # Descrição
        worksheet.set_column(5, 7, 15)  # Valores
        worksheet.set_column(8, 8, 12)  # Estado
        worksheet.set_column(9, 11, 14) # Datas

        # Freeze panes
        worksheet.freeze_panes(6, 0)  # Congela cabeçalhos

        workbook.close()

        # Preparar response
        buffer.seek(0)
        filename = self._gerar_nome_arquivo(filtros, 'xlsx')

        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        return response

    def _gerar_nome_arquivo(self, filtros: Optional[Dict[str, Any]], extensao: str) -> str:
        """
        Gera nome de arquivo baseado nos filtros aplicados

        Args:
            filtros: Dicionário com filtros aplicados
            extensao: Extensão do arquivo (pdf ou xlsx)

        Returns:
            Nome do arquivo
        """
        partes = ['projetos']

        if filtros and isinstance(filtros.get('filters'), dict):
            filters_dict = filtros['filters']

            # Adicionar filtros relevantes
            for key in ['tipo', 'socio', 'estado']:
                if key in filters_dict and filters_dict[key]:
                    valor_sanitizado = str(filters_dict[key]).replace('/', '_').replace(' ', '_')
                    partes.append(f"{key}_{valor_sanitizado}")

        # Adicionar timestamp
        timestamp = self.today.strftime('%Y%m%d')
        partes.append(timestamp)

        return f"{'_'.join(partes)}.{extensao}"


# Funções auxiliares para facilitar o uso
def gerar_relatorio_projetos_pdf(projetos, filtros: Optional[Dict[str, Any]] = None) -> HttpResponse:
    """Atalho para gerar relatório PDF de projetos"""
    relatorio = RelatorioProjetos()
    return relatorio.gerar_pdf(projetos, filtros)


def gerar_relatorio_projetos_excel(projetos, filtros: Optional[Dict[str, Any]] = None) -> HttpResponse:
    """Atalho para gerar relatório Excel de projetos"""
    relatorio = RelatorioProjetos()
    return relatorio.gerar_excel(projetos, filtros)

# ============================================================================
# RELATÓRIOS DE DESPESAS
# ============================================================================

class RelatorioDespesas(RelatorioBase):
    """Geração de relatórios de Despesas"""
    
    def gerar_pdf(self, despesas, filtros: Optional[Dict[str, Any]] = None) -> HttpResponse:
        """Gera relatório PDF de despesas"""
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=1*cm, leftMargin=1*cm,
                                topMargin=1*cm, bottomMargin=1*cm)
        
        elements = []
        styles = getSampleStyleSheet()
        
        # Logo e cabeçalho
        logo_path = os.path.join(settings.BASE_DIR.parent, 'media', 'logos', 'logo_sidebar.png')
        if os.path.exists(logo_path):
            try:
                logo = Image(logo_path, width=3*cm, height=3*cm)
                elements.append(logo)
                elements.append(Spacer(1, 0.2*cm))
            except:
                pass
        
        title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=16, textColor=colors.HexColor('#1976d2'), spaceAfter=3, alignment=TA_CENTER)
        elements.append(Paragraph(self.EMPRESA_MARCA, title_style))
        elements.append(Paragraph(f"{self.EMPRESA_NOME} | NIPC: {self.EMPRESA_NIPC}", styles['Normal']))
        elements.append(Spacer(1, 0.5*cm))
        
        # Título
        titulo = "Relatório de Despesas"
        if filtros:
            if filtros.get('tags'):
                titulo += f" - Tag: {filtros.get('tags')}"
        elements.append(Paragraph(titulo, styles['Heading2']))
        elements.append(Spacer(1, 0.5*cm))
        
        # Tabela
        data = [['#', 'Data', 'Descrição', 'Credor', 'Valor s/ IVA', 'Valor c/ IVA', 'Estado', 'Data Pag.']]
        
        total_sem_iva = Decimal('0.00')
        total_com_iva = Decimal('0.00')
        
        for despesa in despesas:
            descricao = despesa.descricao[:40] + '...' if len(despesa.descricao) > 40 else despesa.descricao
            credor = str(despesa.credor)[:25] + '...' if despesa.credor and len(str(despesa.credor)) > 25 else str(despesa.credor) if despesa.credor else '-'
            
            data.append([
                despesa.numero,
                self._format_date(despesa.data),
                descricao,
                credor,
                self._format_currency(despesa.valor_sem_iva),
                self._format_currency(despesa.valor_com_iva),
                despesa.get_estado_display(),
                self._format_date(despesa.data_pagamento)
            ])
            
            total_sem_iva += despesa.valor_sem_iva or Decimal('0.00')
            total_com_iva += despesa.valor_com_iva or Decimal('0.00')
        
        # Linha de totais
        data.append(['', '', '', 'TOTAL', self._format_currency(total_sem_iva), self._format_currency(total_com_iva), '', ''])
        
        table = Table(data, colWidths=[2*cm, 2.5*cm, 5*cm, 4*cm, 3*cm, 3*cm, 2.5*cm, 2.5*cm])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1976d2')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#e3f2fd')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
        ]))
        
        elements.append(table)
        doc.build(elements)
        
        buffer.seek(0)
        filename = self._gerar_nome_arquivo(filtros, 'pdf')
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    
    def gerar_excel(self, despesas, filtros: Optional[Dict[str, Any]] = None) -> HttpResponse:
        """Gera relatório Excel de despesas"""
        buffer = BytesIO()
        workbook = xlsxwriter.Workbook(buffer, {'in_memory': True})
        worksheet = workbook.add_worksheet('Despesas')
        
        header_format = workbook.add_format({'bold': True, 'bg_color': '#1976d2', 'font_color': 'white', 'align': 'center', 'border': 1})
        currency_format = workbook.add_format({'num_format': '€#,##0.00', 'align': 'right'})
        date_format = workbook.add_format({'num_format': 'dd/mm/yyyy', 'align': 'center'})
        total_format = workbook.add_format({'bold': True, 'bg_color': '#e3f2fd', 'num_format': '€#,##0.00', 'align': 'right'})
        
        headers = ['Número', 'Data', 'Descrição', 'Credor', 'Valor s/ IVA', 'Valor c/ IVA', 'Estado', 'Data Pagamento']
        for col, header in enumerate(headers):
            worksheet.write(0, col, header, header_format)
        
        row = 1
        total_sem_iva = Decimal('0.00')
        total_com_iva = Decimal('0.00')
        
        for despesa in despesas:
            worksheet.write(row, 0, despesa.numero)
            worksheet.write_datetime(row, 1, despesa.data, date_format)
            worksheet.write(row, 2, despesa.descricao)
            worksheet.write(row, 3, str(despesa.credor) if despesa.credor else '-')
            worksheet.write_number(row, 4, float(despesa.valor_sem_iva or 0), currency_format)
            worksheet.write_number(row, 5, float(despesa.valor_com_iva or 0), currency_format)
            worksheet.write(row, 6, despesa.get_estado_display())
            if despesa.data_pagamento:
                worksheet.write_datetime(row, 7, despesa.data_pagamento, date_format)
            
            total_sem_iva += despesa.valor_sem_iva or Decimal('0.00')
            total_com_iva += despesa.valor_com_iva or Decimal('0.00')
            row += 1
        
        worksheet.write(row, 3, 'TOTAL', total_format)
        worksheet.write_number(row, 4, float(total_sem_iva), total_format)
        worksheet.write_number(row, 5, float(total_com_iva), total_format)
        
        workbook.close()
        buffer.seek(0)
        filename = self._gerar_nome_arquivo(filtros, 'xlsx')
        response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    def _gerar_nome_arquivo(self, filtros: Optional[Dict[str, Any]], extensao: str) -> str:
        """Gera nome de arquivo baseado nos filtros aplicados"""
        partes = ['despesas']

        if filtros and isinstance(filtros.get('filters'), dict):
            filters_dict = filtros['filters']
            for key in ['tags', 'estado', 'credor']:
                if key in filters_dict and filters_dict[key]:
                    valor_sanitizado = str(filters_dict[key]).replace('/', '_').replace(' ', '_')
                    partes.append(f"{key}_{valor_sanitizado}")

        timestamp = self.today.strftime('%Y%m%d')
        partes.append(timestamp)

        return f"{'_'.join(partes)}.{extensao}"


def gerar_relatorio_despesas_pdf(despesas, filtros: Optional[Dict[str, Any]] = None) -> HttpResponse:
    relatorio = RelatorioDespesas()
    return relatorio.gerar_pdf(despesas, filtros)


def gerar_relatorio_despesas_excel(despesas, filtros: Optional[Dict[str, Any]] = None) -> HttpResponse:
    relatorio = RelatorioDespesas()
    return relatorio.gerar_excel(despesas, filtros)


# ============================================================================
# RELATÓRIOS DE CLIENTES
# ============================================================================

class RelatorioClientes(RelatorioBase):
    """Geração de relatórios de Clientes"""
    
    def gerar_pdf(self, clientes, filtros: Optional[Dict[str, Any]] = None) -> HttpResponse:
        """Gera relatório PDF de clientes"""
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
        
        elements = []
        styles = getSampleStyleSheet()
        
        # Logo e cabeçalho (continua na próxima linha...)
        logo_path = os.path.join(settings.BASE_DIR.parent, 'media', 'logos', 'logo_sidebar.png')
        if os.path.exists(logo_path):
            try:
                logo = Image(logo_path, width=3*cm, height=3*cm)
                elements.append(logo)
                elements.append(Spacer(1, 0.2*cm))
            except:
                pass
        
        title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=16, textColor=colors.HexColor('#1976d2'), spaceAfter=3, alignment=TA_CENTER)
        elements.append(Paragraph(self.EMPRESA_MARCA, title_style))
        elements.append(Paragraph(f"{self.EMPRESA_NOME} | NIPC: {self.EMPRESA_NIPC}", styles['Normal']))
        elements.append(Spacer(1, 0.5*cm))
        
        titulo = "Relatório de Clientes"
        if filtros and filtros.get('angariador'):
            titulo += f" - Angariador: {filtros.get('angariador')}"
        elements.append(Paragraph(titulo, styles['Heading2']))
        elements.append(Spacer(1, 0.5*cm))
        
        data = [['#', 'Nome', 'NIF', 'Email', 'País', 'Angariador']]
        
        for cliente in clientes:
            nome = cliente.nome[:30] + '...' if len(cliente.nome) > 30 else cliente.nome
            data.append([
                cliente.numero,
                nome,
                cliente.nif or '-',
                cliente.email[:25] + '...' if cliente.email and len(cliente.email) > 25 else cliente.email or '-',
                cliente.pais or '-',
                str(cliente.angariador) if cliente.angariador else '-'
            ])
        
        table = Table(data, colWidths=[2*cm, 5*cm, 3*cm, 4*cm, 2.5*cm, 1.5*cm])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1976d2')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
        ]))
        
        elements.append(table)
        doc.build(elements)
        
        buffer.seek(0)
        filename = self._gerar_nome_arquivo(filtros, 'pdf')
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    
    def gerar_excel(self, clientes, filtros: Optional[Dict[str, Any]] = None) -> HttpResponse:
        """Gera relatório Excel de clientes"""
        buffer = BytesIO()
        workbook = xlsxwriter.Workbook(buffer, {'in_memory': True})
        worksheet = workbook.add_worksheet('Clientes')
        
        header_format = workbook.add_format({'bold': True, 'bg_color': '#1976d2', 'font_color': 'white', 'align': 'center', 'border': 1})
        
        headers = ['Número', 'Nome', 'Nome Formal', 'NIF', 'Email', 'País', 'Contacto', 'Angariador']
        for col, header in enumerate(headers):
            worksheet.write(0, col, header, header_format)
        
        row = 1
        for cliente in clientes:
            worksheet.write(row, 0, cliente.numero)
            worksheet.write(row, 1, cliente.nome)
            worksheet.write(row, 2, cliente.nome_formal)
            worksheet.write(row, 3, cliente.nif or '')
            worksheet.write(row, 4, cliente.email or '')
            worksheet.write(row, 5, cliente.pais or '')
            worksheet.write(row, 6, cliente.contacto or '')
            worksheet.write(row, 7, str(cliente.angariador) if cliente.angariador else '')
            row += 1
        
        workbook.close()
        buffer.seek(0)
        filename = self._gerar_nome_arquivo(filtros, 'xlsx')
        response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    def _gerar_nome_arquivo(self, filtros: Optional[Dict[str, Any]], extensao: str) -> str:
        """Gera nome de arquivo baseado nos filtros aplicados"""
        partes = ['clientes']

        if filtros and isinstance(filtros.get('filters'), dict):
            filters_dict = filtros['filters']
            for key in ['angariador', 'pais']:
                if key in filters_dict and filters_dict[key]:
                    valor_sanitizado = str(filters_dict[key]).replace('/', '_').replace(' ', '_')
                    partes.append(f"{key}_{valor_sanitizado}")

        timestamp = self.today.strftime('%Y%m%d')
        partes.append(timestamp)

        return f"{'_'.join(partes)}.{extensao}"


def gerar_relatorio_clientes_pdf(clientes, filtros: Optional[Dict[str, Any]] = None) -> HttpResponse:
    relatorio = RelatorioClientes()
    return relatorio.gerar_pdf(clientes, filtros)


def gerar_relatorio_clientes_excel(clientes, filtros: Optional[Dict[str, Any]] = None) -> HttpResponse:
    relatorio = RelatorioClientes()
    return relatorio.gerar_excel(clientes, filtros)


# ============================================================================
# RELATÓRIOS DE FORNECEDORES
# ============================================================================

class RelatorioFornecedores(RelatorioBase):
    """Geração de relatórios de Fornecedores"""
    
    def gerar_pdf(self, fornecedores, filtros: Optional[Dict[str, Any]] = None) -> HttpResponse:
        """Gera relatório PDF de fornecedores"""
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=1*cm, leftMargin=1*cm, topMargin=1*cm, bottomMargin=1*cm)
        
        elements = []
        styles = getSampleStyleSheet()
        
        logo_path = os.path.join(settings.BASE_DIR.parent, 'media', 'logos', 'logo_sidebar.png')
        if os.path.exists(logo_path):
            try:
                logo = Image(logo_path, width=3*cm, height=3*cm)
                elements.append(logo)
                elements.append(Spacer(1, 0.2*cm))
            except:
                pass
        
        title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=16, textColor=colors.HexColor('#1976d2'), spaceAfter=3, alignment=TA_CENTER)
        elements.append(Paragraph(self.EMPRESA_MARCA, title_style))
        elements.append(Paragraph(f"{self.EMPRESA_NOME} | NIPC: {self.EMPRESA_NIPC}", styles['Normal']))
        elements.append(Spacer(1, 0.5*cm))
        
        elements.append(Paragraph("Relatório de Fornecedores", styles['Heading2']))
        elements.append(Spacer(1, 0.5*cm))
        
        data = [['#', 'Nome', 'Estatuto', 'Área', 'Função', 'Email', 'NIF', 'País']]
        
        for forn in fornecedores:
            nome = forn.nome[:30] + '...' if len(forn.nome) > 30 else forn.nome
            data.append([
                forn.numero,
                nome,
                forn.get_estatuto_display(),
                forn.area[:20] + '...' if forn.area and len(forn.area) > 20 else forn.area or '-',
                forn.funcao[:20] + '...' if forn.funcao and len(forn.funcao) > 20 else forn.funcao or '-',
                forn.email[:25] + '...' if forn.email and len(forn.email) > 25 else forn.email or '-',
                forn.nif or '-',
                forn.pais or '-'
            ])
        
        table = Table(data, colWidths=[2*cm, 5*cm, 3*cm, 3.5*cm, 3.5*cm, 4.5*cm, 2.5*cm, 2*cm])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1976d2')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
        ]))
        
        elements.append(table)
        doc.build(elements)
        
        buffer.seek(0)
        filename = self._gerar_nome_arquivo(filtros, 'pdf')
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    
    def gerar_excel(self, fornecedores, filtros: Optional[Dict[str, Any]] = None) -> HttpResponse:
        """Gera relatório Excel de fornecedores"""
        buffer = BytesIO()
        workbook = xlsxwriter.Workbook(buffer, {'in_memory': True})
        worksheet = workbook.add_worksheet('Fornecedores')
        
        header_format = workbook.add_format({'bold': True, 'bg_color': '#1976d2', 'font_color': 'white', 'align': 'center', 'border': 1})
        
        headers = ['Número', 'Nome', 'Estatuto', 'Área', 'Função', 'Email', 'NIF', 'País', 'Contacto']
        for col, header in enumerate(headers):
            worksheet.write(0, col, header, header_format)
        
        row = 1
        for forn in fornecedores:
            worksheet.write(row, 0, forn.numero)
            worksheet.write(row, 1, forn.nome)
            worksheet.write(row, 2, forn.get_estatuto_display())
            worksheet.write(row, 3, forn.area or '')
            worksheet.write(row, 4, forn.funcao or '')
            worksheet.write(row, 5, forn.email or '')
            worksheet.write(row, 6, forn.nif or '')
            worksheet.write(row, 7, forn.pais or '')
            worksheet.write(row, 8, forn.contacto or '')
            row += 1
        
        workbook.close()
        buffer.seek(0)
        filename = self._gerar_nome_arquivo(filtros, 'xlsx')
        response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    def _gerar_nome_arquivo(self, filtros: Optional[Dict[str, Any]], extensao: str) -> str:
        """Gera nome de arquivo baseado nos filtros aplicados"""
        partes = ['fornecedores']

        if filtros and isinstance(filtros.get('filters'), dict):
            filters_dict = filtros['filters']
            for key in ['estatuto', 'area', 'pais']:
                if key in filters_dict and filters_dict[key]:
                    valor_sanitizado = str(filters_dict[key]).replace('/', '_').replace(' ', '_')
                    partes.append(f"{key}_{valor_sanitizado}")

        timestamp = self.today.strftime('%Y%m%d')
        partes.append(timestamp)

        return f"{'_'.join(partes)}.{extensao}"


def gerar_relatorio_fornecedores_pdf(fornecedores, filtros: Optional[Dict[str, Any]] = None) -> HttpResponse:
    relatorio = RelatorioFornecedores()
    return relatorio.gerar_pdf(fornecedores, filtros)


def gerar_relatorio_fornecedores_excel(fornecedores, filtros: Optional[Dict[str, Any]] = None) -> HttpResponse:
    relatorio = RelatorioFornecedores()
    return relatorio.gerar_excel(fornecedores, filtros)


# ============================================================================
# RELATÓRIOS DE BOLETINS
# ============================================================================

class RelatorioBoletins(RelatorioBase):
    """Geração de relatórios de Boletins"""
    
    def gerar_pdf(self, boletins, filtros: Optional[Dict[str, Any]] = None) -> HttpResponse:
        """Gera relatório PDF de boletins"""
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
        
        elements = []
        styles = getSampleStyleSheet()
        
        logo_path = os.path.join(settings.BASE_DIR.parent, 'media', 'logos', 'logo_sidebar.png')
        if os.path.exists(logo_path):
            try:
                logo = Image(logo_path, width=3*cm, height=3*cm)
                elements.append(logo)
                elements.append(Spacer(1, 0.2*cm))
            except:
                pass
        
        title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=16, textColor=colors.HexColor('#1976d2'), spaceAfter=3, alignment=TA_CENTER)
        elements.append(Paragraph(self.EMPRESA_MARCA, title_style))
        elements.append(Paragraph(f"{self.EMPRESA_NOME} | NIPC: {self.EMPRESA_NIPC}", styles['Normal']))
        elements.append(Spacer(1, 0.5*cm))
        
        elements.append(Paragraph("Relatório de Boletins", styles['Heading2']))
        elements.append(Spacer(1, 0.5*cm))
        
        data = [['#', 'Mês/Ano', 'Sócio', 'Valor', 'Estado', 'Data Pag.']]
        
        total = Decimal('0.00')
        
        for boletim in boletins:
            data.append([
                boletim.numero,
                f"{boletim.mes}/{boletim.ano}",
                str(boletim.socio),
                self._format_currency(boletim.valor),
                boletim.get_estado_display(),
                self._format_date(boletim.data_pagamento)
            ])
            total += boletim.valor or Decimal('0.00')
        
        data.append(['', '', 'TOTAL', self._format_currency(total), '', ''])
        
        table = Table(data, colWidths=[2*cm, 2.5*cm, 2*cm, 3*cm, 3*cm, 2.5*cm])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1976d2')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#e3f2fd')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
        ]))
        
        elements.append(table)
        doc.build(elements)
        
        buffer.seek(0)
        filename = self._gerar_nome_arquivo(filtros, 'pdf')
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    
    def gerar_excel(self, boletins, filtros: Optional[Dict[str, Any]] = None) -> HttpResponse:
        """Gera relatório Excel de boletins"""
        buffer = BytesIO()
        workbook = xlsxwriter.Workbook(buffer, {'in_memory': True})
        worksheet = workbook.add_worksheet('Boletins')
        
        header_format = workbook.add_format({'bold': True, 'bg_color': '#1976d2', 'font_color': 'white', 'align': 'center', 'border': 1})
        currency_format = workbook.add_format({'num_format': '€#,##0.00', 'align': 'right'})
        date_format = workbook.add_format({'num_format': 'dd/mm/yyyy', 'align': 'center'})
        total_format = workbook.add_format({'bold': True, 'bg_color': '#e3f2fd', 'num_format': '€#,##0.00', 'align': 'right'})
        
        headers = ['Número', 'Mês', 'Ano', 'Sócio', 'Valor', 'Estado', 'Data Pagamento']
        for col, header in enumerate(headers):
            worksheet.write(0, col, header, header_format)
        
        row = 1
        total = Decimal('0.00')
        
        for boletim in boletins:
            worksheet.write(row, 0, boletim.numero)
            worksheet.write(row, 1, boletim.mes)
            worksheet.write(row, 2, boletim.ano)
            worksheet.write(row, 3, str(boletim.socio))
            worksheet.write_number(row, 4, float(boletim.valor or 0), currency_format)
            worksheet.write(row, 5, boletim.get_estado_display())
            if boletim.data_pagamento:
                worksheet.write_datetime(row, 6, boletim.data_pagamento, date_format)
            
            total += boletim.valor or Decimal('0.00')
            row += 1
        
        worksheet.write(row, 3, 'TOTAL', total_format)
        worksheet.write_number(row, 4, float(total), total_format)
        
        workbook.close()
        buffer.seek(0)
        filename = self._gerar_nome_arquivo(filtros, 'xlsx')
        response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-oficedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    def _gerar_nome_arquivo(self, filtros: Optional[Dict[str, Any]], extensao: str) -> str:
        """Gera nome de arquivo baseado nos filtros aplicados"""
        partes = ['boletins']

        if filtros and isinstance(filtros.get('filters'), dict):
            filters_dict = filtros['filters']
            for key in ['socio', 'estado', 'mes', 'ano']:
                if key in filters_dict and filters_dict[key]:
                    valor_sanitizado = str(filters_dict[key]).replace('/', '_').replace(' ', '_')
                    partes.append(f"{key}_{valor_sanitizado}")

        timestamp = self.today.strftime('%Y%m%d')
        partes.append(timestamp)

        return f"{'_'.join(partes)}.{extensao}"


def gerar_relatorio_boletins_pdf(boletins, filtros: Optional[Dict[str, Any]] = None) -> HttpResponse:
    relatorio = RelatorioBoletins()
    return relatorio.gerar_pdf(boletins, filtros)


def gerar_relatorio_boletins_excel(boletins, filtros: Optional[Dict[str, Any]] = None) -> HttpResponse:
    relatorio = RelatorioBoletins()
    return relatorio.gerar_excel(boletins, filtros)


# ============================================================================
# RELATÓRIOS DE ORÇAMENTOS
# ============================================================================

class RelatorioOrcamentos(RelatorioBase):
    """Geração de relatórios de Orçamentos"""
    
    def gerar_pdf(self, orcamentos, filtros: Optional[Dict[str, Any]] = None) -> HttpResponse:
        """Gera relatório PDF de orçamentos"""
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=1*cm, leftMargin=1*cm, topMargin=1*cm, bottomMargin=1*cm)
        
        elements = []
        styles = getSampleStyleSheet()
        
        logo_path = os.path.join(settings.BASE_DIR.parent, 'media', 'logos', 'logo_sidebar.png')
        if os.path.exists(logo_path):
            try:
                logo = Image(logo_path, width=3*cm, height=3*cm)
                elements.append(logo)
                elements.append(Spacer(1, 0.2*cm))
            except:
                pass
        
        title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=16, textColor=colors.HexColor('#1976d2'), spaceAfter=3, alignment=TA_CENTER)
        elements.append(Paragraph(self.EMPRESA_MARCA, title_style))
        elements.append(Paragraph(f"{self.EMPRESA_NOME} | NIPC: {self.EMPRESA_NIPC}", styles['Normal']))
        elements.append(Spacer(1, 0.5*cm))
        
        elements.append(Paragraph("Relatório de Orçamentos", styles['Heading2']))
        elements.append(Spacer(1, 0.5*cm))
        
        data = [['#', 'Cliente', 'Descrição', 'Valor', 'Estado', 'Data Emissão']]
        
        total = Decimal('0.00')
        
        for orc in orcamentos:
            cliente = str(orc.cliente)[:25] + '...' if orc.cliente and len(str(orc.cliente)) > 25 else str(orc.cliente) if orc.cliente else '-'
            descricao = (orc.titulo_cliente or '')[:35] + '...' if orc.titulo_cliente and len(orc.titulo_cliente) > 35 else (orc.titulo_cliente or '-')

            data.append([
                orc.codigo,
                cliente,
                descricao,
                self._format_currency(orc.valor_total),
                orc.get_status_display(),
                self._format_date(orc.data_criacao)
            ])
            total += orc.valor_total or Decimal('0.00')
        
        data.append(['', '', 'TOTAL', self._format_currency(total), '', ''])
        
        table = Table(data, colWidths=[2*cm, 5*cm, 7*cm, 3*cm, 3.5*cm, 3*cm])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1976d2')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#e3f2fd')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
        ]))
        
        elements.append(table)
        doc.build(elements)
        
        buffer.seek(0)
        filename = self._gerar_nome_arquivo(filtros, 'pdf')
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    
    def gerar_excel(self, orcamentos, filtros: Optional[Dict[str, Any]] = None) -> HttpResponse:
        """Gera relatório Excel de orçamentos"""
        buffer = BytesIO()
        workbook = xlsxwriter.Workbook(buffer, {'in_memory': True})
        worksheet = workbook.add_worksheet('Orçamentos')
        
        header_format = workbook.add_format({'bold': True, 'bg_color': '#1976d2', 'font_color': 'white', 'align': 'center', 'border': 1})
        currency_format = workbook.add_format({'num_format': '€#,##0.00', 'align': 'right'})
        date_format = workbook.add_format({'num_format': 'dd/mm/yyyy', 'align': 'center'})
        total_format = workbook.add_format({'bold': True, 'bg_color': '#e3f2fd', 'num_format': '€#,##0.00', 'align': 'right'})
        
        headers = ['Código', 'Cliente', 'Título', 'Valor', 'Status', 'Data Criação']
        for col, header in enumerate(headers):
            worksheet.write(0, col, header, header_format)
        
        row = 1
        total = Decimal('0.00')
        
        for orc in orcamentos:
            worksheet.write(row, 0, orc.codigo)
            worksheet.write(row, 1, str(orc.cliente) if orc.cliente else '')
            worksheet.write(row, 2, orc.titulo_cliente or '')
            worksheet.write_number(row, 3, float(orc.valor_total or 0), currency_format)
            worksheet.write(row, 4, orc.get_status_display())
            worksheet.write_datetime(row, 5, orc.data_criacao, date_format)
            # data_validade não existe no modelo atual

            total += orc.valor_total or Decimal('0.00')
            row += 1
        
        worksheet.write(row, 2, 'TOTAL', total_format)
        worksheet.write_number(row, 3, float(total), total_format)
        
        workbook.close()
        buffer.seek(0)
        filename = self._gerar_nome_arquivo(filtros, 'xlsx')
        response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-oficedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    def _gerar_nome_arquivo(self, filtros: Optional[Dict[str, Any]], extensao: str) -> str:
        """Gera nome de arquivo baseado nos filtros aplicados"""
        partes = ['orcamentos']

        if filtros and isinstance(filtros.get('filters'), dict):
            filters_dict = filtros['filters']
            for key in ['status', 'socio', 'cliente']:
                if key in filters_dict and filters_dict[key]:
                    valor_sanitizado = str(filters_dict[key]).replace('/', '_').replace(' ', '_')
                    partes.append(f"{key}_{valor_sanitizado}")

        timestamp = self.today.strftime('%Y%m%d')
        partes.append(timestamp)

        return f"{'_'.join(partes)}.{extensao}"


def gerar_relatorio_orcamentos_pdf(orcamentos, filtros: Optional[Dict[str, Any]] = None) -> HttpResponse:
    relatorio = RelatorioOrcamentos()
    return relatorio.gerar_pdf(orcamentos, filtros)


def gerar_relatorio_orcamentos_excel(orcamentos, filtros: Optional[Dict[str, Any]] = None) -> HttpResponse:
    relatorio = RelatorioOrcamentos()
    return relatorio.gerar_excel(orcamentos, filtros)
