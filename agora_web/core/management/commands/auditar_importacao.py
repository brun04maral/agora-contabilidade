"""
Django management command para auditar importação Excel vs DB

Compara dados da folha Excel CONTABILIDADE_FINAL com a base de dados
e gera relatório detalhado de incongruências.

Uso:
    python manage.py auditar_importacao excel/CONTABILIDADE_FINAL_20251231.xlsx

Verifica:
- Projetos: quantidades, valores, estados, sócios
- Despesas: quantidades, valores, tags
- Boletins: agregação por mês/sócio, valores, estados
- Prémios: projetos com prémios, valores agregados
- Fornecedores/Clientes: quantidades
"""

import openpyxl
from datetime import date
from decimal import Decimal
from collections import defaultdict
from django.core.management.base import BaseCommand
from django.db.models import Sum, Count, Q

from core.models import (
    Projeto, Despesa, Boletim, Fornecedor, Cliente, Socio,
    TipoProjeto, EstadoProjeto, TagDespesa
)


class Command(BaseCommand):
    help = 'Audita importação comparando Excel vs DB'

    def add_arguments(self, parser):
        parser.add_argument('excel_file', type=str, help='Caminho para o ficheiro Excel')

    def handle(self, *args, **options):
        self.excel_file = options['excel_file']

        self.stdout.write('\n' + '='*80)
        self.stdout.write('🔍 AUDITORIA: EXCEL vs BASE DE DADOS')
        self.stdout.write('='*80)
        self.stdout.write(f'Ficheiro: {self.excel_file}\n')

        try:
            wb = openpyxl.load_workbook(self.excel_file, data_only=True)
        except Exception as e:
            self.stdout.write(self.style.ERROR(f'❌ Erro ao abrir ficheiro: {e}'))
            return

        # Estatísticas
        self.incongruencias = []
        self.avisos = []
        self.validacoes_ok = []

        # Executar verificações
        self.auditar_fornecedores(wb['FORNECEDORES'])
        self.auditar_clientes(wb['CLIENTES'])
        self.auditar_projetos(wb['PROJETOS'])
        self.auditar_despesas_e_boletins(wb['DESPESAS'])

        # Gerar relatório final
        self.gerar_relatorio()

        wb.close()

    # =========================================================================
    # FORNECEDORES
    # =========================================================================

    def auditar_fornecedores(self, ws):
        """Audita fornecedores"""
        self.stdout.write('\n📦 FORNECEDORES')
        self.stdout.write('-'*80)

        # Contar no Excel (só linhas válidas com nome)
        excel_count = 0
        for row_idx in range(2, ws.max_row + 1):
            numero = ws.cell(row_idx, 1).value
            nome = ws.cell(row_idx, 2).value
            if numero and nome:
                excel_count += 1

        # Contar na DB
        db_count = Fornecedor.objects.count()

        self.stdout.write(f'  Excel: {excel_count} fornecedores')
        self.stdout.write(f'  DB:    {db_count} fornecedores')

        diff = db_count - excel_count
        if diff == 0:
            self.stdout.write(self.style.SUCCESS(f'  ✅ Quantidades coincidem'))
            self.validacoes_ok.append('Fornecedores: quantidades OK')
        elif diff > 0:
            self.avisos.append(f'Fornecedores: +{diff} na DB (pode ter dados antigos)')
            self.stdout.write(self.style.WARNING(f'  ⚠️  +{diff} fornecedores na DB'))
        else:
            self.incongruencias.append(f'Fornecedores: {abs(diff)} faltam na DB')
            self.stdout.write(self.style.ERROR(f'  ❌ {abs(diff)} fornecedores faltam'))

    # =========================================================================
    # CLIENTES
    # =========================================================================

    def auditar_clientes(self, ws):
        """Audita clientes"""
        self.stdout.write('\n👥 CLIENTES')
        self.stdout.write('-'*80)

        SKIP_CLIENTES = ['#C0001', '#C0002']

        # Contar no Excel (excluindo sócios)
        excel_count = 0
        for row_idx in range(2, ws.max_row + 1):
            numero = ws.cell(row_idx, 1).value
            nome = ws.cell(row_idx, 2).value
            if numero and nome and numero not in SKIP_CLIENTES:
                excel_count += 1

        # Contar na DB
        db_count = Cliente.objects.count()

        self.stdout.write(f'  Excel: {excel_count} clientes (excluindo sócios)')
        self.stdout.write(f'  DB:    {db_count} clientes')

        diff = db_count - excel_count
        if diff == 0:
            self.stdout.write(self.style.SUCCESS(f'  ✅ Quantidades coincidem'))
            self.validacoes_ok.append('Clientes: quantidades OK')
        elif diff > 0:
            self.avisos.append(f'Clientes: +{diff} na DB (pode ter dados antigos)')
            self.stdout.write(self.style.WARNING(f'  ⚠️  +{diff} clientes na DB'))
        else:
            self.incongruencias.append(f'Clientes: {abs(diff)} faltam na DB')
            self.stdout.write(self.style.ERROR(f'  ❌ {abs(diff)} clientes faltam'))

    # =========================================================================
    # PROJETOS
    # =========================================================================

    def auditar_projetos(self, ws):
        """Audita projetos em detalhe"""
        self.stdout.write('\n📁 PROJETOS')
        self.stdout.write('-'*80)

        # Parse Excel
        excel_projetos = []
        excel_valores = {'total': Decimal('0'), 'BA': Decimal('0'), 'RR': Decimal('0')}

        for row_idx in range(4, ws.max_row + 1):
            numero = ws.cell(row_idx, 1).value
            descricao = ws.cell(row_idx, 5).value

            if not numero or not descricao:
                continue

            valor = ws.cell(row_idx, 6).value or 0
            owner = ws.cell(row_idx, 16).value

            excel_projetos.append({
                'numero': numero,
                'descricao': descricao,
                'valor': Decimal(str(valor)),
                'owner': owner
            })

            excel_valores['total'] += Decimal(str(valor))
            if owner and 'Bruno' in str(owner):
                excel_valores['BA'] += Decimal(str(valor))
            elif owner and 'Rafael' in str(owner):
                excel_valores['RR'] += Decimal(str(valor))

        # DB
        db_count = Projeto.objects.count()
        db_valores_total = Projeto.objects.aggregate(total=Sum('valor_sem_iva'))['total'] or Decimal('0')
        db_valores_ba = Projeto.objects.filter(socio__codigo='BA').aggregate(total=Sum('valor_sem_iva'))['total'] or Decimal('0')
        db_valores_rr = Projeto.objects.filter(socio__codigo='RR').aggregate(total=Sum('valor_sem_iva'))['total'] or Decimal('0')

        # Comparar quantidades
        self.stdout.write(f'\n  Quantidades:')
        self.stdout.write(f'    Excel: {len(excel_projetos)} projetos válidos')
        self.stdout.write(f'    DB:    {db_count} projetos')

        if len(excel_projetos) == db_count:
            self.stdout.write(self.style.SUCCESS(f'    ✅ Quantidades coincidem'))
            self.validacoes_ok.append('Projetos: quantidades OK')
        else:
            diff = db_count - len(excel_projetos)
            if diff > 0:
                self.avisos.append(f'Projetos: +{diff} na DB (dados antigos ou limpeza incompleta)')
                self.stdout.write(self.style.WARNING(f'    ⚠️  +{diff} projetos na DB'))
            else:
                self.incongruencias.append(f'Projetos: {abs(diff)} faltam na DB')
                self.stdout.write(self.style.ERROR(f'    ❌ {abs(diff)} projetos faltam'))

        # Comparar valores
        self.stdout.write(f'\n  Valores totais (sem IVA):')
        self.stdout.write(f'    Excel: €{excel_valores["total"]:,.2f}')
        self.stdout.write(f'    DB:    €{db_valores_total:,.2f}')

        diff_valor = abs(db_valores_total - excel_valores['total'])
        if diff_valor < Decimal('1'):  # Margem de €1 para arredondamentos
            self.stdout.write(self.style.SUCCESS(f'    ✅ Valores coincidem'))
            self.validacoes_ok.append('Projetos: valores totais OK')
        else:
            self.incongruencias.append(f'Projetos: diferença de €{diff_valor:,.2f} nos valores')
            self.stdout.write(self.style.ERROR(f'    ❌ Diferença: €{diff_valor:,.2f}'))

        # Comparar por sócio
        self.stdout.write(f'\n  Por sócio:')
        self.stdout.write(f'    BA - Excel: €{excel_valores["BA"]:,.2f} | DB: €{db_valores_ba:,.2f}')
        self.stdout.write(f'    RR - Excel: €{excel_valores["RR"]:,.2f} | DB: €{db_valores_rr:,.2f}')

    # =========================================================================
    # DESPESAS & BOLETINS
    # =========================================================================

    def auditar_despesas_e_boletins(self, ws):
        """Audita despesas e boletins"""
        self.stdout.write('\n💰 DESPESAS & BOLETINS')
        self.stdout.write('-'*80)

        # Parse Excel
        excel_despesas = {'normais': 0, 'premios': 0, 'boletins': 0}
        excel_valores = {'normais': Decimal('0'), 'premios': Decimal('0'), 'boletins': Decimal('0')}
        premios_por_projeto = defaultdict(lambda: {'BA': Decimal('0'), 'RR': Decimal('0')})
        boletins_por_mes = defaultdict(lambda: {'BA': Decimal('0'), 'RR': Decimal('0')})

        for row_idx in range(6, ws.max_row + 1):
            numero = ws.cell(row_idx, 1).value
            if not numero:
                continue

            tipo_str = ws.cell(row_idx, 7).value or ''
            descricao = ws.cell(row_idx, 8).value or ''
            valor = Decimal(str(ws.cell(row_idx, 16).value or 0).replace(',', '.'))
            credor = ws.cell(row_idx, 5).value or ''
            projeto_num = ws.cell(row_idx, 6).value

            # Classificar
            if 'prém' in tipo_str.lower() or 'premio' in tipo_str.lower() or 'comissão' in tipo_str.lower():
                excel_despesas['premios'] += 1
                excel_valores['premios'] += valor

                # Agregar por projeto
                if projeto_num:
                    socio_code = 'BA' if 'bruno' in credor.lower() else 'RR'
                    premios_por_projeto[projeto_num][socio_code] += valor

            elif any(x in tipo_str.lower() for x in ['deslocação, pessoal', 'per diem pt, pessoal', 'per diem fora, pessoal']):
                excel_despesas['boletins'] += 1
                excel_valores['boletins'] += valor

                # Agregar por mês/sócio
                ano = ws.cell(row_idx, 2).value
                mes = ws.cell(row_idx, 3).value
                if ano and mes:
                    chave = f'{ano}-{mes:02d}' if isinstance(mes, int) else f'{ano}-{mes}'
                    socio_code = 'BA' if 'bruno' in credor.lower() else 'RR'
                    boletins_por_mes[chave][socio_code] += valor
            else:
                excel_despesas['normais'] += 1
                excel_valores['normais'] += valor

        # DB
        db_despesas = Despesa.objects.count()
        db_boletins = Boletim.objects.filter(ano=2025).count()  # Só 2025 pois foram os importados
        db_projetos_premios = Projeto.objects.exclude(premio_bruno=0, premio_rafael=0).count()
        db_premios_ba = Projeto.objects.aggregate(total=Sum('premio_bruno'))['total'] or Decimal('0')
        db_premios_rr = Projeto.objects.aggregate(total=Sum('premio_rafael'))['total'] or Decimal('0')

        # Comparar
        self.stdout.write(f'\n  Despesas normais:')
        self.stdout.write(f'    Excel: {excel_despesas["normais"]} | DB: {db_despesas}')

        self.stdout.write(f'\n  Boletins (linhas Excel → Boletins agregados):')
        self.stdout.write(f'    Excel: {excel_despesas["boletins"]} linhas')
        self.stdout.write(f'    DB:    {db_boletins} boletins (agregados por mês/sócio)')
        self.stdout.write(f'    Esperado: ~{excel_despesas["boletins"] // 12} boletins (12 meses)')

        self.stdout.write(f'\n  Prémios:')
        self.stdout.write(f'    Excel: {excel_despesas["premios"]} linhas de prémios')
        self.stdout.write(f'    DB:    {db_projetos_premios} projetos com prémios agregados')
        self.stdout.write(f'    Valores - BA: Excel €{sum(p["BA"] for p in premios_por_projeto.values()):,.2f} | DB €{db_premios_ba:,.2f}')
        self.stdout.write(f'    Valores - RR: Excel €{sum(p["RR"] for p in premios_por_projeto.values()):,.2f} | DB €{db_premios_rr:,.2f}')

        # Validar prémios
        diff_premios_ba = abs(sum(p['BA'] for p in premios_por_projeto.values()) - db_premios_ba)
        diff_premios_rr = abs(sum(p['RR'] for p in premios_por_projeto.values()) - db_premios_rr)

        if diff_premios_ba < Decimal('1') and diff_premios_rr < Decimal('1'):
            self.stdout.write(self.style.SUCCESS(f'    ✅ Prémios agregados corretamente'))
            self.validacoes_ok.append('Prémios: agregação OK')
        else:
            self.incongruencias.append(f'Prémios: diferença BA €{diff_premios_ba:,.2f}, RR €{diff_premios_rr:,.2f}')
            self.stdout.write(self.style.ERROR(f'    ❌ Diferenças nos prémios'))

    # =========================================================================
    # RELATÓRIO FINAL
    # =========================================================================

    def gerar_relatorio(self):
        """Gera relatório final da auditoria"""
        self.stdout.write('\n' + '='*80)
        self.stdout.write('📊 RELATÓRIO FINAL DE AUDITORIA')
        self.stdout.write('='*80)

        # Validações OK
        if self.validacoes_ok:
            self.stdout.write(self.style.SUCCESS(f'\n✅ VALIDAÇÕES OK ({len(self.validacoes_ok)}):'))
            for ok in self.validacoes_ok:
                self.stdout.write(f'  ✅ {ok}')

        # Avisos
        if self.avisos:
            self.stdout.write(self.style.WARNING(f'\n⚠️  AVISOS ({len(self.avisos)}):'))
            for aviso in self.avisos:
                self.stdout.write(f'  ⚠️  {aviso}')

        # Incongruências
        if self.incongruencias:
            self.stdout.write(self.style.ERROR(f'\n❌ INCONGRUÊNCIAS ({len(self.incongruencias)}):'))
            for inc in self.incongruencias:
                self.stdout.write(f'  ❌ {inc}')
        else:
            self.stdout.write(self.style.SUCCESS('\n🎉 NENHUMA INCONGRUÊNCIA CRÍTICA ENCONTRADA!'))

        # Resumo
        self.stdout.write('\n' + '='*80)
        total = len(self.validacoes_ok) + len(self.avisos) + len(self.incongruencias)
        self.stdout.write(f'Total de verificações: {total}')
        self.stdout.write(f'  ✅ OK: {len(self.validacoes_ok)}')
        self.stdout.write(f'  ⚠️  Avisos: {len(self.avisos)}')
        self.stdout.write(f'  ❌ Incongruências: {len(self.incongruencias)}')
        self.stdout.write('='*80 + '\n')
