"""
Django management command para importar dados da Google Sheet (Excel export)

Uso:
    python manage.py import_from_excel excel/CONTABILIDADE_FINAL_20251231.xlsx

Este command importa dados das seguintes abas:
- FORNECEDORES → Model Fornecedor
- CLIENTES → Model Cliente (skip sócios #C0001, #C0002)
- PROJETOS → Model Projeto
- DESPESAS → Model Despesa + Boletim (conversão)

Ver documentação completa em: docs/EXCEL_IMPORT_ANALYSIS.md
"""

import openpyxl
from datetime import date, datetime
from decimal import Decimal
from collections import defaultdict
from django.core.management.base import BaseCommand
from django.db import transaction
from core.models import (
    Socio, Cliente, Fornecedor, Projeto, Despesa, Boletim,
    TagDespesa, TipoProjeto, EstadoProjeto, EstadoBoletim,
    EstatutoFornecedor
)


class Command(BaseCommand):
    help = 'Importa dados da Google Sheet (Excel export)'

    def add_arguments(self, parser):
        parser.add_argument('excel_file', type=str, help='Caminho para o ficheiro Excel')
        parser.add_argument(
            '--dry-run',
            action='store_true',
            help='Simula import sem gravar na BD'
        )
        parser.add_argument(
            '--skip-fornecedores',
            action='store_true',
            help='Não importar Fornecedores'
        )
        parser.add_argument(
            '--skip-clientes',
            action='store_true',
            help='Não importar Clientes'
        )
        parser.add_argument(
            '--skip-projetos',
            action='store_true',
            help='Não importar Projetos'
        )
        parser.add_argument(
            '--skip-despesas',
            action='store_true',
            help='Não importar Despesas/Boletins'
        )

    def handle(self, *args, **options):
        self.dry_run = options['dry_run']
        self.excel_file = options['excel_file']

        if self.dry_run:
            self.stdout.write(self.style.WARNING('🔍 DRY RUN MODE - Nenhuma alteração será gravada'))

        self.stdout.write(f'\n📊 A importar de: {self.excel_file}\n')

        try:
            wb = openpyxl.load_workbook(self.excel_file, data_only=True)
        except Exception as e:
            self.stdout.write(self.style.ERROR(f'❌ Erro ao abrir ficheiro: {e}'))
            return

        # Estatísticas
        self.stats = {
            'fornecedores': 0,
            'clientes': 0,
            'projetos': 0,
            'despesas': 0,
            'boletins': 0,
            'premios_agregados': 0,
            'erros': []
        }

        try:
            with transaction.atomic():
                # 1. Fornecedores
                if not options['skip_fornecedores']:
                    self.stdout.write('\n1️⃣  Importando FORNECEDORES...')
                    self.import_fornecedores(wb['FORNECEDORES'])

                # 2. Clientes
                if not options['skip_clientes']:
                    self.stdout.write('\n2️⃣  Importando CLIENTES...')
                    self.import_clientes(wb['CLIENTES'])

                # 3. Projetos
                if not options['skip_projetos']:
                    self.stdout.write('\n3️⃣  Importando PROJETOS...')
                    self.import_projetos(wb['PROJETOS'])

                # 4. Despesas (3 fases)
                if not options['skip_despesas']:
                    self.stdout.write('\n4️⃣  Processando DESPESAS...')
                    self.process_despesas(wb['DESPESAS'])

                if self.dry_run:
                    raise Exception("DRY RUN - rollback transaction")

        except Exception as e:
            if not self.dry_run:
                self.stdout.write(self.style.ERROR(f'\n❌ Erro durante import: {e}'))
                raise

        # Resumo
        self.print_summary()

        wb.close()

    # =========================================================================
    # FORNECEDORES
    # =========================================================================

    def import_fornecedores(self, ws):
        """Importa fornecedores da aba FORNECEDORES"""
        header_row = 1

        for row_idx in range(header_row + 1, ws.max_row + 1):
            try:
                numero = ws.cell(row_idx, 1).value
                if not numero:
                    continue

                nome = ws.cell(row_idx, 2).value
                estatuto_str = ws.cell(row_idx, 3).value
                area = ws.cell(row_idx, 4).value
                funcao = ws.cell(row_idx, 5).value
                classificacao_str = ws.cell(row_idx, 6).value
                validade_seguro = ws.cell(row_idx, 7).value
                nif = ws.cell(row_idx, 8).value
                iban = ws.cell(row_idx, 9).value
                morada = ws.cell(row_idx, 10).value
                pais = ws.cell(row_idx, 11).value or 'Portugal'
                contacto = ws.cell(row_idx, 12).value
                email = ws.cell(row_idx, 13).value
                nota = ws.cell(row_idx, 14).value

                # Parse estatuto
                estatuto = self.parse_estatuto_fornecedor(estatuto_str)

                # Parse classificação (estrelas)
                classificacao = self.parse_classificacao(classificacao_str)

                # Parse NIF (pode vir como float)
                if nif:
                    nif = str(int(float(nif)))

                Fornecedor.objects.update_or_create(
                    numero=numero,
                    defaults={
                        'nome': nome,
                        'estatuto': estatuto,
                        'area': area or '',
                        'funcao': funcao or '',
                        'classificacao': classificacao,
                        'validade_seguro_trabalho': validade_seguro,
                        'nif': nif,
                        'iban': iban,
                        'morada': morada,
                        'pais': pais,
                        'contacto': contacto,
                        'email': email,
                        'nota': nota,
                    }
                )
                self.stats['fornecedores'] += 1

            except Exception as e:
                self.stats['erros'].append(f'Fornecedor linha {row_idx}: {e}')

        self.stdout.write(self.style.SUCCESS(f'   ✅ {self.stats["fornecedores"]} fornecedores'))

    # =========================================================================
    # CLIENTES
    # =========================================================================

    SKIP_CLIENTES = ['#C0001', '#C0002']  # Sócios

    def import_clientes(self, ws):
        """Importa clientes da aba CLIENTES (skip sócios)"""
        header_row = 1

        for row_idx in range(header_row + 1, ws.max_row + 1):
            try:
                numero = ws.cell(row_idx, 1).value
                if not numero or numero in self.SKIP_CLIENTES:
                    continue

                nome = ws.cell(row_idx, 2).value
                nif = ws.cell(row_idx, 3).value
                morada = ws.cell(row_idx, 4).value
                pais = ws.cell(row_idx, 5).value or 'Portugal'
                angariacao = ws.cell(row_idx, 6).value
                nota = ws.cell(row_idx, 8).value

                # Parse NIF
                if nif:
                    nif = str(int(float(nif)))

                Cliente.objects.update_or_create(
                    numero=numero,
                    defaults={
                        'nome': nome,
                        'nome_formal': nome,  # Sheet não diferencia
                        'nif': nif,
                        'morada': morada,
                        'pais': pais,
                        'angariacao': angariacao,
                        'nota': nota,
                    }
                )
                self.stats['clientes'] += 1

            except Exception as e:
                self.stats['erros'].append(f'Cliente linha {row_idx}: {e}')

        self.stdout.write(self.style.SUCCESS(f'   ✅ {self.stats["clientes"]} clientes'))

    # =========================================================================
    # PROJETOS
    # =========================================================================

    def import_projetos(self, ws):
        """Importa projetos da aba PROJETOS"""
        header_row = 3

        self.stdout.write(f'   🔍 DEBUG: Aba PROJETOS tem {ws.max_row} linhas')

        for row_idx in range(header_row + 1, ws.max_row + 1):
            try:
                numero = ws.cell(row_idx, 1).value
                if not numero:
                    continue

                cliente_nome = ws.cell(row_idx, 2).value
                data_inicio = ws.cell(row_idx, 3).value
                data_fim = ws.cell(row_idx, 4).value
                descricao = ws.cell(row_idx, 5).value
                valor_sem_iva = ws.cell(row_idx, 6).value or 0
                data_faturacao = ws.cell(row_idx, 7).value
                data_vencimento = ws.cell(row_idx, 8).value
                data_recibo = ws.cell(row_idx, 9).value
                orcamento_url = ws.cell(row_idx, 10).value
                equipa = ws.cell(row_idx, 11).value
                recursos_humanos = ws.cell(row_idx, 12).value
                equipamento_usado = ws.cell(row_idx, 13).value
                local = ws.cell(row_idx, 14).value
                estado_sheet = ws.cell(row_idx, 15).value
                owner_nome = ws.cell(row_idx, 16).value
                nota = ws.cell(row_idx, 17).value

                # Lookups
                cliente = self.lookup_cliente(cliente_nome) if cliente_nome else None
                socio = self.lookup_socio(owner_nome) if owner_nome else None

                # Determinar tipo e estado
                tipo, estado = self.parse_projeto_tipo_estado(
                    estado_sheet,
                    data_recibo,
                    data_fim
                )

                Projeto.objects.update_or_create(
                    numero=numero,
                    defaults={
                        'tipo': tipo,
                        'estado': estado,
                        'cliente': cliente,
                        'socio': socio,
                        'owner': socio.codigo if socio else 'BA',  # Deprecated
                        'data_inicio': self.parse_date(data_inicio),
                        'data_fim': self.parse_date(data_fim),
                        'descricao': descricao or '',
                        'valor_sem_iva': Decimal(str(valor_sem_iva)),
                        'data_faturacao': self.parse_date(data_faturacao),
                        'data_vencimento': self.parse_date(data_vencimento),
                        'data_recibo': self.parse_date(data_recibo),
                        'orcamento_url': orcamento_url,
                        'equipa': int(equipa) if equipa else None,
                        'recursos_humanos': recursos_humanos,
                        'equipamento_usado': equipamento_usado,
                        'local': local,
                        'nota': nota,
                    }
                )
                self.stats['projetos'] += 1

            except Exception as e:
                self.stats['erros'].append(f'Projeto linha {row_idx}: {e}')

        self.stdout.write(self.style.SUCCESS(f'   ✅ {self.stats["projetos"]} projetos'))

    # =========================================================================
    # DESPESAS (3 fases)
    # =========================================================================

    def process_despesas(self, ws):
        """Processa despesas em 3 fases: prémios, boletins, normais"""

        # Parse todas as despesas
        despesas_raw = self.parse_all_despesas(ws)

        # Fase 1: Agregar prémios
        self.stdout.write('   📊 Fase 1: Agregando prémios...')
        self.aggregate_premios(despesas_raw['premios'])

        # Fase 2: Criar boletins
        self.stdout.write('   📊 Fase 2: Criando boletins...')
        self.create_boletins(despesas_raw['boletins'])

        # Fase 3: Importar despesas normais
        self.stdout.write('   📊 Fase 3: Importando despesas...')
        self.import_despesas(despesas_raw['normais'])

        self.stdout.write(self.style.SUCCESS(
            f'   ✅ {self.stats["despesas"]} despesas, '
            f'{self.stats["boletins"]} boletins, '
            f'{self.stats["premios_agregados"]} prémios agregados'
        ))

    def parse_all_despesas(self, ws):
        """Parse todas as despesas e separa em 3 categorias"""
        premios = []
        boletins_raw = []
        normais = []

        header_row = 5

        for row_idx in range(header_row + 1, ws.max_row + 1):
            numero = ws.cell(row_idx, 1).value
            if not numero:
                continue

            tipo_str = ws.cell(row_idx, 7).value or ''
            tags = self.parse_tags(tipo_str)

            despesa_raw = {
                'numero': numero,
                'ano': int(ws.cell(row_idx, 2).value or 0),
                'mes': int(ws.cell(row_idx, 3).value or 0),
                'dia': int(ws.cell(row_idx, 4).value or 0),
                'credor_nome': ws.cell(row_idx, 5).value,
                'projeto_numero': ws.cell(row_idx, 6).value,
                'tipo_original': tipo_str,
                'tags': tags,
                'descricao': ws.cell(row_idx, 8).value or '',
                'valor_sem_iva': Decimal(str(ws.cell(row_idx, 10).value or 0)),
                'valor_com_iva': Decimal(str(ws.cell(row_idx, 13).value or 0)),
                'data_vencimento': ws.cell(row_idx, 20).value,
                'nota': ws.cell(row_idx, 23).value,
            }

            # Classificar
            if 'PREMIO' in tags or 'COMISSAO_VENDA' in tags:
                premios.append(despesa_raw)
            elif any(t in tags for t in ['PER_DIEM_PT', 'PER_DIEM_FORA']) or \
                 ('DESLOCACAO' in tags and 'PESSOAL' in tags):
                boletins_raw.append(despesa_raw)
            else:
                normais.append(despesa_raw)

        return {
            'premios': premios,
            'boletins': boletins_raw,
            'normais': normais
        }

    def aggregate_premios(self, premios):
        """Agrega prémios por projeto e popula premio_bruno/premio_rafael"""
        premios_por_projeto = defaultdict(lambda: {'BA': Decimal('0'), 'RR': Decimal('0')})

        for premio in premios:
            projeto_numero = premio['projeto_numero']
            if not projeto_numero:
                continue

            socio = self.identify_socio_from_credor(premio['credor_nome'])
            if not socio:
                continue

            premios_por_projeto[projeto_numero][socio.codigo] += premio['valor_sem_iva']

        # Atualizar projetos
        for projeto_numero, valores in premios_por_projeto.items():
            try:
                Projeto.objects.filter(numero=projeto_numero).update(
                    premio_bruno=valores['BA'],
                    premio_rafael=valores['RR']
                )
                self.stats['premios_agregados'] += 1
            except Exception as e:
                self.stats['erros'].append(f'Erro ao agregar prémio {projeto_numero}: {e}')

    def create_boletins(self, boletins_raw):
        """Cria boletins agrupados por (socio, mes, ano)"""
        grupos = defaultdict(list)

        # Agrupar
        for desp in boletins_raw:
            socio = self.identify_socio_from_credor(desp['credor_nome'])
            if not socio:
                continue

            chave = (socio.codigo, desp['ano'], desp['mes'])
            grupos[chave].append(desp)

        # Criar boletins
        for (socio_codigo, ano, mes), despesas in grupos.items():
            try:
                socio = Socio.objects.get(codigo=socio_codigo)

                # Verificar estado
                datas_venc = [d['data_vencimento'] for d in despesas if d['data_vencimento']]
                estado = EstadoBoletim.PAGO if datas_venc else EstadoBoletim.PENDENTE
                data_pag = max(datas_venc) if datas_venc else None

                # Criar boletim
                Boletim.objects.update_or_create(
                    socio=socio,
                    mes=mes,
                    ano=ano,
                    defaults={
                        'descricao': self.get_month_name(mes),
                        'data_emissao': date(ano, mes, 27),  # Fixo dia 27
                        'data_pagamento': self.parse_date(data_pag) if data_pag else None,
                        'valor_total': sum(d['valor_sem_iva'] for d in despesas),
                        'estado': estado,
                    }
                )
                self.stats['boletins'] += 1

            except Exception as e:
                self.stats['erros'].append(f'Erro ao criar boletim {socio_codigo} {mes}/{ano}: {e}')

    def import_despesas(self, despesas_raw):
        """Importa despesas normais com tags"""
        for desp in despesas_raw:
            try:
                # Lookups
                credor = self.lookup_fornecedor(desp['credor_nome']) if desp['credor_nome'] else None
                projeto = self.lookup_projeto(desp['projeto_numero']) if desp['projeto_numero'] else None

                # Criar data
                data_desp = date(desp['ano'], desp['mes'], desp['dia']) if all([desp['ano'], desp['mes'], desp['dia']]) else date.today()

                # Criar despesa
                despesa = Despesa.objects.create(
                    numero=desp['numero'],
                    data=data_desp,
                    credor=credor,
                    projeto=projeto,
                    descricao=desp['descricao'],
                    valor_sem_iva=desp['valor_sem_iva'],
                    valor_com_iva=desp['valor_com_iva'],
                    tipo_original=desp['tipo_original'],
                    estado='PAGO' if desp['data_vencimento'] else 'PENDENTE',
                    data_pagamento=self.parse_date(desp['data_vencimento']) if desp['data_vencimento'] else None,
                    nota=desp['nota'],
                )

                # Adicionar tags
                for tag_codigo in desp['tags']:
                    try:
                        tag = TagDespesa.objects.get(codigo=tag_codigo)
                        despesa.tags.add(tag)
                    except TagDespesa.DoesNotExist:
                        self.stats['erros'].append(f'Tag {tag_codigo} não existe')

                self.stats['despesas'] += 1

            except Exception as e:
                self.stats['erros'].append(f'Despesa {desp.get("numero", "???")}: {e}')

    # =========================================================================
    # HELPER METHODS
    # =========================================================================

    def parse_estatuto_fornecedor(self, estatuto_str):
        """Parse estatuto do fornecedor"""
        if not estatuto_str:
            return EstatutoFornecedor.FREELANCER

        estatuto_map = {
            'EMPRESA': EstatutoFornecedor.EMPRESA,
            'FREELANCER': EstatutoFornecedor.FREELANCER,
            'ESTADO': EstatutoFornecedor.ESTADO,
            'BANCO': EstatutoFornecedor.BANCO,
            'SÓCIO GERENTE': EstatutoFornecedor.SOCIO_GERENTE,
        }

        return estatuto_map.get(estatuto_str.upper(), EstatutoFornecedor.FREELANCER)

    def parse_classificacao(self, stars_str):
        """Parse classificação de estrelas"""
        if not stars_str:
            return None
        return stars_str.count('★') or stars_str.count('*')

    def parse_tags(self, tipo_str):
        """Parse tipo composto em tags"""
        if not tipo_str:
            return []

        parts = [p.strip() for p in tipo_str.split(',')]

        mapping = {
            'Prémio': 'PREMIO',
            'Serviço': 'SERVICO',
            'Equipamento': 'EQUIPAMENTO',
            'Pessoal': 'PESSOAL',
            'Comissão venda': 'COMISSAO_VENDA',
            'Administrativo': 'ADMINISTRATIVO',
            'Ordenado': 'ORDENADO',
            'Sub. Alimentação': 'SUB_ALIMENTACAO',
            'Alimentação': 'ALIMENTACAO',
            'Produção': 'PRODUCAO',
            'Deslocação': 'DESLOCACAO',
            'Per Diem PT': 'PER_DIEM_PT',
            'Per Diem FORA': 'PER_DIEM_FORA',
            'IRS Retenção': 'IRS_RETENCAO',
        }

        tags = []
        for part in parts:
            if part in mapping:
                tags.append(mapping[part])

        return tags

    def parse_projeto_tipo_estado(self, estado_sheet, data_recibo, data_fim):
        """Determina tipo e estado do projeto"""
        # Tipo
        if estado_sheet == 'Pessoal':
            tipo = TipoProjeto.PESSOAL
        else:
            tipo = TipoProjeto.EMPRESA

        # Estado
        if data_recibo:
            estado = EstadoProjeto.PAGO
        elif estado_sheet == 'Pessoal':
            if data_fim and self.parse_date(data_fim) < date.today():
                estado = EstadoProjeto.FINALIZADO
            else:
                estado = EstadoProjeto.ATIVO
        elif estado_sheet == 'Finalizado':
            estado = EstadoProjeto.FINALIZADO
        else:
            estado = EstadoProjeto.ATIVO

        return tipo, estado

    def lookup_cliente(self, nome):
        """Lookup cliente por nome"""
        try:
            return Cliente.objects.filter(nome__icontains=nome).first()
        except:
            return None

    def lookup_fornecedor(self, nome):
        """Lookup fornecedor por nome"""
        try:
            return Fornecedor.objects.filter(nome__icontains=nome).first()
        except:
            return None

    def lookup_projeto(self, numero):
        """Lookup projeto por número"""
        try:
            return Projeto.objects.get(numero=numero)
        except:
            return None

    def lookup_socio(self, nome):
        """Lookup sócio por nome"""
        if not nome:
            return None

        if 'Bruno' in nome or 'Amaral' in nome:
            return Socio.objects.get(codigo='BA')
        elif 'Rafael' in nome or 'Reigota' in nome:
            return Socio.objects.get(codigo='RR')

        return None

    def identify_socio_from_credor(self, credor_nome):
        """Identifica sócio a partir do nome do credor"""
        return self.lookup_socio(credor_nome)

    def parse_date(self, value):
        """Parse date from Excel"""
        if not value:
            return None
        # Check datetime FIRST (datetime is subclass of date!)
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        return None

    def get_month_name(self, mes):
        """Retorna nome do mês em português"""
        meses = [
            '', 'Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho',
            'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro'
        ]
        return meses[mes] if 1 <= mes <= 12 else str(mes)

    def print_summary(self):
        """Imprime resumo final"""
        self.stdout.write('\n' + '='*80)
        self.stdout.write(self.style.SUCCESS('✅ IMPORT CONCLUÍDO!'))
        self.stdout.write('='*80)
        self.stdout.write(f'📦 Fornecedores: {self.stats["fornecedores"]}')
        self.stdout.write(f'👥 Clientes: {self.stats["clientes"]}')
        self.stdout.write(f'📁 Projetos: {self.stats["projetos"]}')
        self.stdout.write(f'💰 Despesas: {self.stats["despesas"]}')
        self.stdout.write(f'📋 Boletins: {self.stats["boletins"]}')
        self.stdout.write(f'🏆 Prémios agregados: {self.stats["premios_agregados"]}')

        if self.stats['erros']:
            self.stdout.write(f'\n⚠️  {len(self.stats["erros"])} erros')

            # Erros de projetos (primeiros 10)
            erros_projetos = [e for e in self.stats['erros'] if 'Projeto' in e]
            if erros_projetos:
                self.stdout.write(f'\n   🔴 Erros de PROJETOS ({len(erros_projetos)} total, primeiros 10):')
                for erro in erros_projetos[:10]:
                    self.stdout.write(self.style.WARNING(f'      - {erro}'))

            # Outros erros (primeiros 10)
            outros_erros = [e for e in self.stats['erros'] if 'Projeto' not in e]
            if outros_erros:
                self.stdout.write(f'\n   🟡 Outros erros ({len(outros_erros)} total, primeiros 10):')
                for erro in outros_erros[:10]:
                    self.stdout.write(self.style.WARNING(f'      - {erro}'))

        self.stdout.write('\n' + '='*80 + '\n')
