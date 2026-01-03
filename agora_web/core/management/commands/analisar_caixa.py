"""
Django management command para analisar aba CAIXA do Excel

Extrai fórmulas, analisa lógica de cálculo, compara com SaldosCalculator
e gera documentação completa.

Uso:
    python manage.py analisar_caixa excel/CONTABILIDADE_FINAL_20251231.xlsx [--output docs/CAIXA_ANALYSIS.md]

Funcionalidades:
- Extrai todas as fórmulas da aba CAIXA
- Interpreta a lógica de negócio
- Compara valores Excel vs SaldosCalculator
- Gera documentação em Markdown
"""

import openpyxl
import re
from decimal import Decimal
from django.core.management.base import BaseCommand
from django.db.models import Sum, Q

from core.models import Projeto, Despesa, Boletim, Socio
from core.utils.saldos import SaldosCalculator


class Command(BaseCommand):
    help = 'Analisa aba CAIXA e compara com SaldosCalculator'

    def add_arguments(self, parser):
        parser.add_argument('excel_file', type=str, help='Caminho para o ficheiro Excel')
        parser.add_argument(
            '--output',
            type=str,
            default=None,
            help='Ficheiro Markdown para gravar documentação'
        )

    def handle(self, *args, **options):
        self.excel_file = options['excel_file']
        self.output_file = options['output']

        self.stdout.write('\n' + '='*80)
        self.stdout.write('🔍 ANÁLISE DETALHADA DA ABA CAIXA')
        self.stdout.write('='*80)
        self.stdout.write(f'Excel: {self.excel_file}\n')

        try:
            # Abrir Excel com fórmulas
            wb = openpyxl.load_workbook(self.excel_file, data_only=False)
            ws_formulas = wb['CAIXA']

            # Abrir Excel com valores
            wb_valores = openpyxl.load_workbook(self.excel_file, data_only=True)
            ws_valores = wb_valores['CAIXA']

        except Exception as e:
            self.stdout.write(self.style.ERROR(f'❌ Erro ao abrir ficheiro: {e}'))
            return

        # Executar análises
        self.analisar_estrutura(ws_formulas, ws_valores)
        self.extrair_formulas(ws_formulas, ws_valores)
        self.comparar_com_saldos_calculator(ws_valores)
        self.gerar_documentacao(ws_formulas, ws_valores)

        wb.close()
        wb_valores.close()

        self.stdout.write('\n' + '='*80)
        self.stdout.write(self.style.SUCCESS('✅ Análise concluída!'))
        self.stdout.write('='*80 + '\n')

    # =========================================================================
    # ANÁLISE DE ESTRUTURA
    # =========================================================================

    def analisar_estrutura(self, ws_formulas, ws_valores):
        """Analisa a estrutura da aba CAIXA"""
        self.stdout.write('\n📊 ESTRUTURA DA ABA CAIXA')
        self.stdout.write('-'*80)

        # Identificar headers
        headers = {}
        for col_idx in range(1, 15):
            cell = ws_valores.cell(2, col_idx)
            if cell.value:
                headers[col_idx] = str(cell.value)

        self.stdout.write('\n  Headers encontrados:')
        for col_idx, header in headers.items():
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            self.stdout.write(f'    {col_letter}: {header}')

        # Identificar linhas de dados
        self.stdout.write('\n  Linhas de dados:')
        for row_idx in range(3, 12):
            label = ws_valores.cell(row_idx, 2).value
            if label:
                self.stdout.write(f'    Linha {row_idx}: {label}')

    # =========================================================================
    # EXTRAÇÃO DE FÓRMULAS
    # =========================================================================

    def extrair_formulas(self, ws_formulas, ws_valores):
        """Extrai e interpreta fórmulas principais"""
        self.stdout.write('\n📝 FÓRMULAS PRINCIPAIS')
        self.stdout.write('-'*80)

        # Células chave para analisar
        celulas_chave = {
            'C4': 'Investimento Bruno',
            'D4': 'Prémios Bruno (Excel)',
            'E4': 'Projetos Pessoais Bruno (Excel)',
            'F4': 'Prémios não faturados Bruno (Excel)',
            'G4': 'Total INs Bruno (Excel)',
            'H4': 'Despesas Fixas Bruno (Excel)',
            'I4': 'Boletins Bruno (Excel)',
            'J4': 'Despesas Pessoais Bruno (Excel)',
            'C5': 'Investimento Rafael',
            'D5': 'Prémios Rafael (Excel)',
            'E5': 'Projetos Pessoais Rafael (Excel)',
            'F5': 'Prémios não faturados Rafael (Excel)',
            'G5': 'Total INs Rafael (Excel)',
            'H5': 'Despesas Fixas Rafael (Excel)',
            'I5': 'Boletins Rafael (Excel)',
            'J5': 'Despesas Pessoais Rafael (Excel)',
        }

        for coord, descricao in celulas_chave.items():
            formula = ws_formulas[coord].value
            valor = ws_valores[coord].value

            self.stdout.write(f'\n  {coord} - {descricao}')
            self.stdout.write(f'    Valor: {valor}')

            if formula and isinstance(formula, str) and formula.startswith('='):
                # Simplificar fórmula para exibição
                formula_clean = self.simplificar_formula(formula)
                self.stdout.write(f'    Fórmula: {formula_clean}')

    def simplificar_formula(self, formula):
        """Simplifica fórmula para exibição"""
        if len(formula) > 100:
            # Remover __xludf.DUMMYFUNCTION e IFERROR externos
            formula = re.sub(r'__xludf\.DUMMYFUNCTION\("', '', formula)
            formula = re.sub(r'IFERROR\(', '', formula)
            # Truncar se muito longa
            if len(formula) > 150:
                return formula[:147] + '...'
        return formula

    # =========================================================================
    # COMPARAÇÃO COM SALDOSCALCULATOR
    # =========================================================================

    def comparar_com_saldos_calculator(self, ws_valores):
        """Compara valores Excel com SaldosCalculator"""
        self.stdout.write('\n💰 COMPARAÇÃO: EXCEL vs SALDOSCALCULATOR')
        self.stdout.write('-'*80)

        # Calcular saldos usando SaldosCalculator
        calc = SaldosCalculator()
        saldos_ba = calc.calcular_saldo_bruno()
        saldos_rr = calc.calcular_saldo_rafael()

        # Extrair valores do Excel
        excel_ba = {
            'investimento': self._get_decimal(ws_valores, 'C4'),
            'premios': self._get_decimal(ws_valores, 'D4'),
            'projetos_pessoais': self._get_decimal(ws_valores, 'E4'),
            'premios_nao_faturados': self._get_decimal(ws_valores, 'F4'),
            'total_ins': self._get_decimal(ws_valores, 'G4'),
            'despesas_fixas': self._get_decimal(ws_valores, 'H4'),
            'boletins': self._get_decimal(ws_valores, 'I4'),
            'despesas_pessoais': self._get_decimal(ws_valores, 'J4'),
        }

        excel_rr = {
            'investimento': self._get_decimal(ws_valores, 'C5'),
            'premios': self._get_decimal(ws_valores, 'D5'),
            'projetos_pessoais': self._get_decimal(ws_valores, 'E5'),
            'premios_nao_faturados': self._get_decimal(ws_valores, 'F5'),
            'total_ins': self._get_decimal(ws_valores, 'G5'),
            'despesas_fixas': self._get_decimal(ws_valores, 'H5'),
            'boletins': self._get_decimal(ws_valores, 'I5'),
            'despesas_pessoais': self._get_decimal(ws_valores, 'J5'),
        }

        # Comparar Bruno
        self.stdout.write('\n  👤 BRUNO (BA):')
        self._comparar_socio('BA', excel_ba, saldos_ba)

        # Comparar Rafael
        self.stdout.write('\n  👤 RAFAEL (RR):')
        self._comparar_socio('RR', excel_rr, saldos_rr)

    def _comparar_socio(self, codigo, excel_vals, calc_vals):
        """Compara valores de um sócio"""
        comparacoes = [
            ('Projetos Pessoais', excel_vals['projetos_pessoais'], calc_vals['ins'].get('projetos_pessoais', 0)),
            ('Prémios', excel_vals['premios'], calc_vals['ins'].get('premios', 0)),
            ('Despesas Fixas', excel_vals['despesas_fixas'], calc_vals['outs'].get('despesas_fixas', 0)),
            ('Boletins', excel_vals['boletins'], calc_vals['outs'].get('boletins_total', 0)),
        ]

        for label, excel_val, calc_val in comparacoes:
            excel_val = excel_val or Decimal('0')
            calc_val = Decimal(str(calc_val))
            diff = abs(excel_val - calc_val)

            status = '✅' if diff < Decimal('1') else '❌'
            self.stdout.write(f'    {status} {label}:')
            self.stdout.write(f'       Excel: €{excel_val:,.2f}')
            self.stdout.write(f'       Calc:  €{calc_val:,.2f}')
            if diff >= Decimal('1'):
                self.stdout.write(f'       Diff:  €{diff:,.2f}')

    def _get_decimal(self, ws, coord):
        """Obtém valor decimal de uma célula"""
        val = ws[coord].value
        if val is None:
            return Decimal('0')
        return Decimal(str(val))

    # =========================================================================
    # GERAÇÃO DE DOCUMENTAÇÃO
    # =========================================================================

    def gerar_documentacao(self, ws_formulas, ws_valores):
        """Gera documentação em Markdown"""
        self.stdout.write('\n📄 DOCUMENTAÇÃO')
        self.stdout.write('-'*80)

        if not self.output_file:
            self.stdout.write('  ⚠️  Sem ficheiro de output especificado (use --output)')
            return

        doc_lines = []
        doc_lines.append('# Análise da Aba CAIXA - Lógica de Saldos Pessoais')
        doc_lines.append('')
        doc_lines.append('**Data:** 03 Janeiro 2026')
        doc_lines.append('**Fonte:** CONTABILIDADE_FINAL_20251231.xlsx')
        doc_lines.append('')
        doc_lines.append('---')
        doc_lines.append('')
        doc_lines.append('## 📊 Estrutura da Aba')
        doc_lines.append('')
        doc_lines.append('A aba CAIXA contém o cálculo de saldos pessoais dos sócios.')
        doc_lines.append('')
        doc_lines.append('### Headers (Linha 2)')
        doc_lines.append('')

        # Headers
        for col_idx in range(2, 15):
            header = ws_valores.cell(2, col_idx).value
            if header:
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                doc_lines.append(f'- **{col_letter}**: {header}')

        doc_lines.append('')
        doc_lines.append('### Linhas de Dados')
        doc_lines.append('')
        doc_lines.append('| Linha | Sócio | Descrição |')
        doc_lines.append('|-------|-------|-----------|')
        doc_lines.append('| 4 | BA | Bruno Amaral |')
        doc_lines.append('| 5 | RR | Rafael Reigota |')
        doc_lines.append('')
        doc_lines.append('---')
        doc_lines.append('')
        doc_lines.append('## 💡 Fórmulas Principais')
        doc_lines.append('')

        # Fórmulas por categoria
        categorias = {
            'INs (Entradas - Empresa DEVE ao sócio)': [
                ('C4/C5', 'Investimento Inicial', True),
                ('D4/D5', 'Prémios (pagos)', False),
                ('E4/E5', 'Projetos Pessoais (pagos)', False),
                ('F4/F5', 'Prémios não faturados', False),
                ('G4/G5', 'TOTAL INs', True),
            ],
            'OUTs (Saídas - Empresa PAGOU ao sócio)': [
                ('H4/H5', 'Despesas Fixas Mensais', False),
                ('I4/I5', 'Boletins (ajudas de custo)', False),
                ('J4/J5', 'Despesas Pessoais', False),
            ],
        }

        for categoria, formulas in categorias.items():
            doc_lines.append(f'### {categoria}')
            doc_lines.append('')

            for coord_template, descricao, is_static in formulas:
                # Usar linha 4 (BA) como exemplo
                coord = coord_template.split('/')[0]
                formula = ws_formulas[coord].value
                valor = ws_valores[coord].value

                doc_lines.append(f'#### {descricao} ({coord})')
                doc_lines.append('')
                doc_lines.append(f'**Valor (BA):** €{valor:,.2f}' if valor else '**Valor (BA):** -')
                doc_lines.append('')

                if formula and isinstance(formula, str) and formula.startswith('='):
                    if is_static:
                        doc_lines.append(f'**Fórmula:** Valor fixo')
                    else:
                        # Tentar interpretar a fórmula
                        interpretacao = self._interpretar_formula(formula)
                        if interpretacao:
                            doc_lines.append(f'**Lógica:** {interpretacao}')
                        doc_lines.append('')
                        doc_lines.append('```excel')
                        doc_lines.append(formula)
                        doc_lines.append('```')
                doc_lines.append('')

        doc_lines.append('---')
        doc_lines.append('')
        doc_lines.append('## 🔍 Comparação com SaldosCalculator')
        doc_lines.append('')
        doc_lines.append('Ver comando: `python manage.py analisar_caixa`')
        doc_lines.append('')

        # Gravar ficheiro
        try:
            with open(self.output_file, 'w', encoding='utf-8') as f:
                f.write('\n'.join(doc_lines))
            self.stdout.write(self.style.SUCCESS(f'  ✅ Documentação gravada em: {self.output_file}'))
        except Exception as e:
            self.stdout.write(self.style.ERROR(f'  ❌ Erro ao gravar: {e}'))

    def _interpretar_formula(self, formula):
        """Tenta interpretar a lógica de uma fórmula"""
        if 'SUMIFS' in formula and 'PROJETOS' in formula and 'pessoal' in formula:
            return 'Soma projetos pessoais PAGOS'
        elif 'SUMIFS' in formula and 'DESPESAS' in formula and 'Mensal' in formula:
            return 'Soma despesas fixas mensais ÷ 2'
        elif 'FILTER' in formula and 'DESPESAS' in formula and 'Prémio' in formula:
            return 'Filtra e soma prémios pagos'
        elif 'SUM' in formula and ('C4' in formula or 'C5' in formula):
            return 'Soma todos os INs'
        return None
